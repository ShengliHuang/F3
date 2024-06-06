import glob
import numpy as np
import numpy.ma as ma
import os, sys, traceback, datetime, time
from datetime import date
import shutil
import math
import csv
import openpyxl   #added on 20240513 for processing excel XLSX files
import sqlite3
import multiprocessing
from subprocess import Popen   #Popen is used for running BAT file in python
import pandas as pd

#print("Is it possible to run EXE file in Linux? yes, see https://www.wikihow.com/Can-Linux-Run-Exe. This means when using docker/container, we do not have to use windows container!")
#print("DRAST need post-fire plots, but we may use FVS [mortality modifier] to simulate the burned plot with some information from RAVG such as burned severity and basal area and crown closure % loss")

F3MetricExcelXLSXfile = r'F:\CUI\fhaastf3app\F3DataEveryWhere\F3MetricInformation.xlsx'
FIA_FVS_SpeciesCrosswalk = r'F:\CUI\fhaastf3app\FVS\FVS_SpeciesCrosswalk.xlsx'   #20240206, the format must be xlsx, so I convert the original excel format to xlsx
Tiles = ["MichiganTile3"]#["MichiganTile1","MichiganTile2","MichiganTile3","MichiganTile4","MichiganTile5"]#,"MichiganTile3"]   #This can be automatically but I prefer manual input
FVSINPUTDBpath = r'F:\CUI\fhaastf3app\FIA'  
FVSprocessingpath = r'F:\CUI\fhaastf3app\FVS'  
RemoteSensingYearOfInterest = [2023]
F3Log = os.getcwd()+os.sep+"FVSLog.txt"
F3LogWrongFVS = os.getcwd()+os.sep+"FVSLog_PlotIsWrong.txt"
Factor = -11.1
t0 = time.time()






   

def FIAspeciesListFromFVSINPUTDBsBasedOnPlotlist(FVSINPUTDBs,plotlist):
    try:
        print("This function may take some time, but it is worth, as we do not need to manually type the species for FVS run")
        plotlist = [str(k) for k in plotlist]
        FiaSpecieslist = []
        for FVSINPUTDB in FVSINPUTDBs:
            print("Finding the FIA species for the tile based on FIA survey tree list from ",FVSINPUTDB)
            conn = sqlite3.connect(FVSINPUTDB)  #https://www.sqlitetutorial.net/sqlite-python/
            cur= conn.cursor()
            cur.execute("SELECT * FROM FVS_TREEINIT_PLOT")  #Ask Marcus to give the table name consistently
            columns = [column[0] for column in cur.description]
            #print(columns)
            STAND_IDIndex = columns.index("STAND_ID")
            #print("STAND_IDIndex=",STAND_IDIndex)
            SPECIESIndex = columns.index("SPECIES")   #In FVSINPUTDB (i.e.,FVS data in FIADB), the "SPECIES" refers to FIA species code
            #print("SPECIESIndex=",SPECIESIndex)
            rows = cur.fetchall()

            rowid = 0
            for row in rows:
                rowid = rowid + 1
                STAND_ID = str(row[STAND_IDIndex])
                SPECIES = str(row[SPECIESIndex])
                FIADB_PLOT = CreateFIADBUniquePlotIDfromFvsReadyStandID(STAND_ID)  
                #print(rowid,FIADB_PLOT)   #The FIADB tree table is very big, we do not want to print each line in the screen. 
                if FIADB_PLOT in plotlist:
                    FiaSpecieslist.append(SPECIES)
        print("The number of FIA surveyed trees in this HSL are: ",len(FiaSpecieslist))
        FiaSpecieslistUnique = sorted(set(FiaSpecieslist))  #unique and sort, see https://stackoverflow.com/questions/2931672/what-is-the-cleanest-way-to-do-a-sort-plus-uniq-on-a-python-list
        print("The final FIA species code unqiue in this HSL file are: ",FiaSpecieslistUnique)
        return FiaSpecieslistUnique
    except Exception as e:
        print("Sorry, there is an error : ",e)
        exc_type, exc_value, exc_traceback = sys.exc_info()
        strTrace = traceback.format_exception(exc_type, exc_value, exc_traceback)
        ErrorLog = os.getcwd()+os.sep+"FvsError.txt"
        ErrorLogFile = open(ErrorLog, 'a')
        Content1 = "Dr. Shengli Huang detected an error at "+datetime.datetime.now().strftime("%Y-%m-%d %H:%M:%S") + "\n"
        #print(Content1)
        Content2 = "The error was for FIAspeciesListFromFVSINPUTDBsBasedOnPlotlist with inputs of " + repr(FVSINPUTDBs)+" "+ repr(plotlist)+ " with the following details:\n"
        #print(Content2)
        Content3 = repr(strTrace) + "\n\n\n"  #see https://docs.python.org/2/library/traceback.html
        #print(Content3)
        ErrorLogFile.write(Content1)
        ErrorLogFile.write(Content2)
        ErrorLogFile.write(Content3)
        ErrorLogFile.close()
        return ErrorLog


def FVSalphaCodeCrosswalkedFromFIAspecies(FIAspeciesList,FIA_FVS_SpeciesCrosswalk,SpeciesTranslatorInThisHSL):
    try:
        FIAspeciesList = [int(k) for k in FIAspeciesList]
        print(FIAspeciesList)
        WestOrEasternFVSvariant = SpeciesTranslatorInThisHSL
        print("WestOrEasternFVSvariant=",WestOrEasternFVSvariant)
        SpeciesCrosswalkDF = pd.read_excel(FIA_FVS_SpeciesCrosswalk,WestOrEasternFVSvariant)  #https://www.geeksforgeeks.org/reading-excel-file-using-python/
        #print("SpeciesCrosswalkDF.dtypes=",SpeciesCrosswalkDF.dtypes) 
        #print("The first several records are:", SpeciesCrosswalkDF.head())
        SpeciesCrosswalkDFList = SpeciesCrosswalkDF.values.tolist()
        print("The number of record of ",FIA_FVS_SpeciesCrosswalk," is:", len(SpeciesCrosswalkDFList)," and the first record is:",SpeciesCrosswalkDFList[0])

        FvsAlphaSpecies = []
        for FIAspecies in FIAspeciesList:
            #print(FIAspecies)
            FVSALPHACODEFOUND = "No"
            for k in range(0,len(SpeciesCrosswalkDFList),1):
                ThisRecord = SpeciesCrosswalkDFList[k]
                FIACODE = ThisRecord[2]  #fields in the FVS_SpeciesCrosswalk.xlsx are: Unique Identifier,USDA PLANTS SYMBOL,FIA CODE,FVS ALPHA CODE,COMMON NAME,SCIENTIFIC NAME,CS_Mapped_To,LS_Mapped_To,NE_Mapped_To,SN_Mapped_To
                if not math.isnan(FIACODE):   
                    if int(FIACODE) == FIAspecies:
                        FVSALPHACODE = ThisRecord[3]
                        print("FIACODE =",str(int(FIACODE)).zfill(3)," corresponds to FVS Alpha Code of ",FVSALPHACODE)
                        FvsAlphaSpecies.append(FVSALPHACODE)
                        FVSALPHACODEFOUND = "Yes"
            if FVSALPHACODEFOUND == "No":
                FvsAlphaSpecies.append("NotFound") 
        return FvsAlphaSpecies
    except Exception as e:
        print("Sorry, there is an error : ",e)
        exc_type, exc_value, exc_traceback = sys.exc_info()
        strTrace = traceback.format_exception(exc_type, exc_value, exc_traceback)
        ErrorLog = os.getcwd()+os.sep+"FvsError.txt"
        ErrorLogFile = open(ErrorLog, 'a')
        Content1 = "Dr. Shengli Huang detected an error at "+datetime.datetime.now().strftime("%Y-%m-%d %H:%M:%S") + "\n"
        #print(Content1)
        Content2 = "The error was for FVSalphaCodeCrosswalkedFromFIAspecies with inputs of " + repr(FIAspeciesList)+" "+ repr(FIA_FVS_SpeciesCrosswalk)+" "+repr(SpeciesTranslatorInThisHSL) + " with the following details:\n"
        #print(Content2)
        Content3 = repr(strTrace) + "\n\n\n"  #see https://docs.python.org/2/library/traceback.html
        #print(Content3)
        ErrorLogFile.write(Content1)
        ErrorLogFile.write(Content2)
        ErrorLogFile.write(Content3)
        ErrorLogFile.close()
        return ErrorLog



def PlotID_FIAspeciesCode_FVSalphaSpecies(ShengliHuangKeyFile,Factor):
    try:
        ShengliHuangKey = open(ShengliHuangKeyFile, "rb")
        MyBinaryLine = ShengliHuangKey.readlines()  #readline only read one line
        MyBinaryLine = [i.decode().replace('\n','') for i in MyBinaryLine]  #We had error [A bytes-like object is required, not 'str'], so We must use decode() here, see https://stackoverflow.com/questions/50829364/cannot-split-a-bytes-like-object-is-required-not-str
        SpeciesCodeAlreadyInThisHSL = "NotFound"
        plotlist = []
        for line in MyBinaryLine:
            if line.startswith('###'):
                if "The FIA species code in this tile are" in line:
                    print("FIAspeciesListInThisHSL=",line)
                    FIAspeciesListInThisHSL = line.split("The FIA species code in this tile are:")[1].split(",")
                    SpeciesCodeAlreadyInThisHSL = "Found" 
                if "The FVS alpha code in this tile are" in line:
                    print("FVSalphaCodeListInThisHSL=",line)
                    FVSalphaCodeListInThisHSL = line.split("The FVS alpha code in this tile are:")[1].split(",")
            if not line.startswith('###'):  #[###] was added in the file on 20240202
                plot = str(round(float(line.split(",")[0]) / float(Factor)))  #20230509: change int() to round(). e.g. round(2.999)=3 while int(2.999)=2. also 20240212: add str() to convert the plot into string not integer
                x = float(line.split(",")[1]) / float(Factor)
                y = float(line.split(",")[2]) / float(Factor)
                plotlist.append(str(plot)) #20240212: plot change to str(plot)
        ShengliHuangKey.close()
        if SpeciesCodeAlreadyInThisHSL == "Found":
            FVSalphaCodeListInThisHSLReturned = [FVSalphaCodeListInThisHSL[k] for k in range(0,len(FVSalphaCodeListInThisHSL),1) if FVSalphaCodeListInThisHSL[k] != "NotFound"]
            FIAspeciesListInThisHSLReturned = [FIAspeciesListInThisHSL[k] for k in range(0,len(FVSalphaCodeListInThisHSL),1) if FVSalphaCodeListInThisHSL[k] != "NotFound"]
            return plotlist,FIAspeciesListInThisHSLReturned,FVSalphaCodeListInThisHSLReturned
        if SpeciesCodeAlreadyInThisHSL == "NotFound":
            plotlist,xlist,ylist,StatesCodeInThisHSL,StatesFullNameInThisHSL,StatesShortNameInThisHSL,SpeciesTranslatorInThisHSL = PlotIDandXandY(ShengliHuangKeyFile,Factor)
            FVSINPUTDBs = [FVSINPUTDBpath+os.sep+"SQLite_FIADB_"+k+".db" for k in StatesShortNameInThisHSL]
            #print("FVSINPUTDBs=",FVSINPUTDBs)
            FIAspeciesListInThisHSL = FIAspeciesListFromFVSINPUTDBsBasedOnPlotlist(FVSINPUTDBs,plotlist)  #FIAspeciesListFromFVSINPUTDBs returns a list of FIAspecies
            print("FIAspeciesListInThisHSL=",FIAspeciesListInThisHSL)        
            FVSalphaCodeListInThisHSL = FVSalphaCodeCrosswalkedFromFIAspecies(FIAspeciesListInThisHSL,FIA_FVS_SpeciesCrosswalk,SpeciesTranslatorInThisHSL)  #FVSalphaCodeCrosswalkedFromFIAspecies return a list of FVSalphaCode
            FIAspeciesListInThisHSL = [k.zfill(3) for k in FIAspeciesListInThisHSL]  #This convert FIAspecies to three digits. Note THIS SENTENCE MUST BE AFTER FVSalphaCodeCrosswalkedFromFIAspecies() sentence above
            #print("11111FVSalphaCodeListInThisHSL=",FVSalphaCodeListInThisHSL)
            FVSalphaCodeListInThisHSL = ["NotFound" if x is np.nan else x for x in FVSalphaCodeListInThisHSL]  #Very important
            #print("222222FVSalphaCodeListInThisHSL=",FVSalphaCodeListInThisHSL)
            FVSalphaCodeListInThisHSLReturned = [FVSalphaCodeListInThisHSL[k] for k in range(0,len(FVSalphaCodeListInThisHSL),1) if FVSalphaCodeListInThisHSL[k] != "NotFound"]
            FIAspeciesListInThisHSLReturned = [FIAspeciesListInThisHSL[k] for k in range(0,len(FVSalphaCodeListInThisHSL),1) if FVSalphaCodeListInThisHSL[k] != "NotFound"]
            #20240207: Add this section to reduce the processing time in the future use. Also can be used for converting FIA species and FVS species at the end of F3 processing---start
            #check https://stackoverflow.com/questions/11229780/adding-text-and-lines-to-the-beginning-of-a-file-python

            f = open(ShengliHuangKeyFile,"rb")
            lines = f.readlines() # read old content
            lines = [i.decode().replace('\n','') for i in lines]  #We had error [A bytes-like object is required, not 'str'], so We must use decode() here, see https://stackoverflow.com/questions/50829364/cannot-split-a-bytes-like-object-is-required-not-str
            f.close()
            ShengliHuangKeyFileImproved = ShengliHuangKeyFile.replace(".hsl","_Improved.hsl")
            f1 = open(ShengliHuangKeyFileImproved,"wb")
            LineID = 0
            for line in lines: # write old content after new
                LineID = LineID + 1
                line = line + "\n"
                f1.write(line.encode('utf-8'))
                if LineID == 5: #Because there are already 5 line with ###
                    FIAspeciescodeLine = "###The FIA species code in this tile are:"+",".join(FIAspeciesListInThisHSL)+"\n"
                    f1.write(FIAspeciescodeLine.encode('utf-8')) # write new content at the beginning
                    FVSalphaCodeLine = "###The FVS alpha code in this tile are:"+",".join(FVSalphaCodeListInThisHSL)+"\n"
                    f1.write(FVSalphaCodeLine.encode('utf-8')) # write new content at the beginning   
            f1.close()
            os.remove(ShengliHuangKeyFile)
            os.rename(ShengliHuangKeyFileImproved, ShengliHuangKeyFile)
            #20240207: Add this section to reduce the processing time in the future use. Also can be used for converting FIA species and FVS species at the end of F3 processing---end
            return plotlist,FIAspeciesListInThisHSLReturned,FVSalphaCodeListInThisHSLReturned
    except Exception as e:
        print("Sorry, there is an error : ",e)
        exc_type, exc_value, exc_traceback = sys.exc_info()
        strTrace = traceback.format_exception(exc_type, exc_value, exc_traceback)
        ErrorLog = os.getcwd()+os.sep+"FvsError.txt"
        ErrorLogFile = open(ErrorLog, 'a')
        Content1 = "Dr. Shengli Huang detected an error at "+datetime.datetime.now().strftime("%Y-%m-%d %H:%M:%S") + "\n"
        #print(Content1)
        Content2 = "The error was for PlotID_FIAspeciesCode_FVSalphaSpecies with inputs of " + repr(ShengliHuangKeyFile)+" "+ repr(Factor)+ " with the following details:\n"
        #print(Content2)
        Content3 = repr(strTrace) + "\n\n\n"  #see https://docs.python.org/2/library/traceback.html
        #print(Content3)
        ErrorLogFile.write(Content1)
        ErrorLogFile.write(Content2)
        ErrorLogFile.write(Content3)
        ErrorLogFile.close()
        return ErrorLog




def PlotIDandXandY(ShengliHuangKeyFile,Factor):
    try:
        ShengliHuangKey = open(ShengliHuangKeyFile, "rb")
        MyBinaryLine = ShengliHuangKey.readlines()  #readline only read one line
        MyBinaryLine = [i.decode().replace('\n','') for i in MyBinaryLine]  #We had error [A bytes-like object is required, not 'str'], so We must use decode() here, see https://stackoverflow.com/questions/50829364/cannot-split-a-bytes-like-object-is-required-not-str
        plotlist = []
        xlist = []
        ylist = []
        for line in MyBinaryLine:
            if line.startswith('###'):
                if "The states Code (from FVS) spatially covered" in line:
                    StatesCodeInThisHSL = line.split(" include ")[1].split(",")  #Noth this sentence with the word [ include ] is very important
                if "The states full name spatially covered" in line:
                    StatesFullNameInThisHSL = line.split(" include ")[1].split(",")  #Noth this sentence with the word [ include ] is very important
                if "The states short name spatially covered" in line:
                    StatesShortNameInThisHSL = line.split(" include ")[1].split(",")  #Noth this sentence with the word [ include ] is very important
                if "FVS easter variants (CS, LS, NE, SN) and western variants (AK, BM, CA, CI, CR, EC, EM, IE, KT, NC, OC, OP, PN, SO, TT, UT, WC, WS) have" in line:
                    SpeciesTranslatorInThisHSL = line.split("This tile uses ")[1]  #Noth this sentence with the word [This tile use ] is very important
            if not line.startswith('###'):  #[###] was added in the file on 20240202
                plot = round(float(line.split(",")[0]) / float(Factor))  #20230509: change int() to round(). e.g. round(2.999)=3 while int(2.999)=2
                x = float(line.split(",")[1]) / float(Factor)
                y = float(line.split(",")[2]) / float(Factor)
                plotlist.append(str(plot))  #20240212: plot changed to str(plot)
                xlist.append(x)
                ylist.append(y)
        ShengliHuangKey.close()
        return plotlist,xlist,ylist,StatesCodeInThisHSL,StatesFullNameInThisHSL,StatesShortNameInThisHSL,SpeciesTranslatorInThisHSL
    except Exception as e:
        print("Sorry, there is an error : ",e)
        exc_type, exc_value, exc_traceback = sys.exc_info()
        strTrace = traceback.format_exception(exc_type, exc_value, exc_traceback)
        ErrorLog = os.getcwd()+os.sep+"F3Error.txt"
        ErrorLogFile = open(ErrorLog, 'a')
        Content1 = "Dr. Shengli Huang detected an error at "+datetime.datetime.now().strftime("%Y-%m-%d %H:%M:%S") + "\n"  
        #print(Content1)    
        Content2 = "The error was for PlotIDandXandY with inputs of " + repr(ShengliHuangKeyFile)+ repr(Factor)+" with the following details:\n"
        #print(Content2)
        Content3 = repr(strTrace) + "\n\n\n"  #see https://docs.python.org/2/library/traceback.html
        #print(Content3)
        ErrorLogFile.write(Content1)
        ErrorLogFile.write(Content2)
        ErrorLogFile.write(Content3)
        ErrorLogFile.close()
        return ErrorLog


def SpeciesSizeSpecficMetricsKcp(SpeciesList, SizeList, MetricList, SpeciesGroup):
    try:
        #SPMCDBH(measurement, species, tree-value class, lower DBH, upper DBH, lower height, upper height, tree status, point number)
        #First argument: 1 = trees per acre 2 = basal area per acre 3 = total cubic foot volume/acre for western variants; total merchantable cubic foot volume/acre (pulpwood + sawtimber)
        #for eastern variants 4 = merchantable board foot volume/acre for western variants; merchantable sawtimber board foot volume/acre for eastern variants 5 = quadratic mean diameter
        #6 = average height 7 = percent cover (corrected for crown overlap, based on stockable area) 8 = average dwarf mistletoe rating
        #9 = merchantable cubic foot volume/acre for western variants; merchantable cubic foot sawtimber volume/acre for eastern variants
        #10 = average diameter growth for the cycle 11 = stand density index (Reineke 1933) (based on stockable area)
        #12 = relative density as defined in SILVAH (Marquis and Ernst 1984) (based on stockable area)
        #13 = stand density index (Zeide 1983) (based on stockable area)   14 = Curtis’ relative density index (Curtis 2010) (based on stockable area)
        #Note: The maximum number of columns is 32767 in a sqlite table, which should be sufficient for our F3

        SpeciesSizeSpecficMetricsKcp = FVSprocessingpath + os.sep + "GeneralKcpKey" + os.sep + Tile + os.sep + "SpeciesSizeSpecficMetrics.kcp"
        if os.path.exists(SpeciesSizeSpecficMetricsKcp):
            today = date.today()
            epoch_time = int(time.time())   #https://stackoverflow.com/questions/4548684/how-to-get-the-seconds-since-epoch-from-the-time-date-output-of-gmtime
            SpeciesSizeSpecficMetricsKcpBackup = FVSprocessingpath + os.sep + "GeneralKcpKey" + os.sep + Tile + os.sep + "SpeciesSizeSpecficMetrics_epochtime"+str(epoch_time)+".kcp"   #SpeciesSizeSpecficMetricsKcpBackup = os.getcwd() + os.sep + "SpeciesSizeSpecficMetrics_Backup"+str(today)+".kcp"   
            os.rename(SpeciesSizeSpecficMetricsKcp, SpeciesSizeSpecficMetricsKcpBackup)
        SpeciesSizeSpecficMetricsKcpFile = open(SpeciesSizeSpecficMetricsKcp,mode='w')  #https://www.tutorialspoint.com/how-to-read-a-text-file-in-python
        Comment0 = "*This FVS KCP for species-specific and size-specific variables was created by Dr. Shengli Huang (shengli.huang@usda.gov) at "+datetime.datetime.now().strftime("%Y-%m-%d %H:%M:%S") 
        Comment1 = "*There are "+str(len(SpeciesList))+" species and they are: " + ",".join(SpeciesList)
        Comment2 = "*There are "+str(len(SizeList))+" size classes and they are: " + ",".join(SizeList)
        Comment3 = "*There are "+str(len(MetricList))+" metrics and they are: " + ",".join(MetricList)
        Comment4 = "*SPGROUP can be used to define up to 10 species group, with each group containing up to 50 species. If sentence is >150 characters, use & to combine. Also note 1) SPGROUP should not be within a compute block, 2) be careful of the space in [SpGroup       ConOak]"
        Comment5 = "*SpGroup       ConOak"
        Comment6 = "*BO NP WO BR CK RO SW SM PB WP BF TA EH JP RN"
        Comment7 = "Here ConOak is the group name while [BO NP WO BR CK RO SW SM PB WP BF TA EH JP RN] are the species alpha code included in this group"
        SpeciesSizeSpecficMetricsKcpFile.write(Comment0+"\n")
        SpeciesSizeSpecficMetricsKcpFile.write(Comment1+"\n")
        SpeciesSizeSpecficMetricsKcpFile.write(Comment2+"\n")
        SpeciesSizeSpecficMetricsKcpFile.write(Comment3+"\n")
        SpeciesSizeSpecficMetricsKcpFile.write(Comment4+"\n")
        SpeciesSizeSpecficMetricsKcpFile.write(Comment5+"\n")
        SpeciesSizeSpecficMetricsKcpFile.write(Comment6+"\n")
        SpeciesSizeSpecficMetricsKcpFile.write(Comment7+"\n\n")

        for k in SpeciesGroup:
            SpeciesGroupLine1 = "SpGroup       "+ k.split(":")[0] + "\n"   #Note the space length after SpGroup is very important
            SpeciesSizeSpecficMetricsKcpFile.write(SpeciesGroupLine1)
            SpeciesList.append(k.split(":")[0])  #k.split(":")[0] is the speciesgroup name
            SpeciesGroupLine2 = " ".join(k.split(":")[1].split(",")) + "\n\n"
            SpeciesSizeSpecficMetricsKcpFile.write(SpeciesGroupLine2)

        SpeciesSizeSpecficMetricsKcpFile.write("FMIN"+"\n")
        SpeciesSizeSpecficMetricsKcpFile.write("End"+"\n")
        SpeciesSizeSpecficMetricsKcpFile.write("Compute            0"+"\n")
        BASA_VariableNameList = []
        BASA_SizeList = []
        BASA_SpeciesList = []
        SpeciesID = 0
        for Species in SpeciesList:
            SpeciesID = SpeciesID + 1
            SizeID = 0
            for Size in SizeList:
                SizeID = SizeID + 1
                MetricID = 0
                for Metric in MetricList:
                    MetricID = MetricID + 1
                    VariableName = "S"+str(MetricID)+str(SpeciesID)+str(SizeID)  #Here we use S to represent SPMCDBH function (i.e., the first letter)
                    #VariableName = Metric+str(SpeciesID)+str(SizeID)
                    VariableName1 = Metric+"_"+Species+"_"+Size.replace("-","_")  #20240125: I found SQLITE does not like "-" in the field name
                    CommmentLine = "* The FVS only allows maximum 8 charaters, so variable name of " + VariableName + " below is actually " + VariableName1
                    SpeciesSizeSpecficMetricsKcpFile.write(CommmentLine+"\n")
                    if Metric == "TPA":
                        SPMCDBH_argument1 = "1"
                    if Metric == "BASA":
                        SPMCDBH_argument1 = "2"
                    if Metric == "TCFVOL":
                        SPMCDBH_argument1 = "3"
                    if Metric == "MBFVOL":
                        SPMCDBH_argument1 = "4"
                    if Metric == "QMD":
                        SPMCDBH_argument1 = "5"
                    if Metric == "AHGT":
                        SPMCDBH_argument1 = "6"
                    if Metric == "CORPC":
                        SPMCDBH_argument1 = "7"
                    if Metric == "MISTLE":
                        SPMCDBH_argument1 = "8"
                    if Metric == "MCFVOL":
                        SPMCDBH_argument1 = "9"
                    if Metric == "AVGDBHGROW":
                        SPMCDBH_argument1 = "10"
                    if Metric == "SDI1933":
                        SPMCDBH_argument1 = "11"
                    if Metric == "RELDEN1984":
                        SPMCDBH_argument1 = "12"
                    if Metric == "SDI1983":
                        SPMCDBH_argument1 = "13"
                    if Metric == "RELDEN2010":
                        SPMCDBH_argument1 = "14"
                    if Species.lower() == "all":
                        SPMCDBH_argument2 = "0"
                    else:
                        #print("It is very important to give the correct short name of species for this variant. Check FVS mannual please")
                        SPMCDBH_argument2 = Species  #Species code (alpha code, species sequence number, species group name, or species group sequence number). Only trees of the specified species or species group will be included. Code a zero (or All) for all species. Numeric species codes are variant specific.
                    SPMCDBH_argument3 = "0"   #Tree-value class (IMC codes 1,2, or 3; see section 4.2.1). Code a zero for all value classes
                    SPMCDBH_argument4 = Size.split("_")[0]
                    SPMCDBH_argument5 = Size.split("_")[1]
                    SPMCDBH_argument6 = "0"
                    SPMCDBH_argument7 = "999"
                    SPMCDBH_argument8 = "0"  #0 = live trees (this is the default value), 
                    SPMCDBH_argument9 = "0"
                    SPMCDBH_ArgumentList = [SPMCDBH_argument1,SPMCDBH_argument2,SPMCDBH_argument3,SPMCDBH_argument4,SPMCDBH_argument5,SPMCDBH_argument6,SPMCDBH_argument7,SPMCDBH_argument8,SPMCDBH_argument9]
                    ThisLine = VariableName + " = SPMCDBH(" + ",".join(SPMCDBH_ArgumentList) + ")"
                    SpeciesSizeSpecficMetricsKcpFile.write(ThisLine+"\n")
                    if "BASA" in Metric:
                        if "0,999,0,999" in ThisLine:  #20240131 added: The species absence 0 and presnece 0 is not used for DBH size and height size. Instead, it refers to the DBH 0f 0-999 and height of 0-999 (i.e., include all)
                            BASA_VariableNameList.append(VariableName)
                            BASA_SizeList.append(Size)
                            BASA_SpeciesList.append(Species)
        #In addition to SPMCDBH(), we can add more functions such as TREEBIO(), snags() etc in this function
        SpeciesSizeSpecficMetricsKcpFile.write("End"+"\n")

        #202401424: This is to calculate the species 0 absence and 1 presence based on basal area value. The purpose is to impute the species confidence requested by Marla---start
        #[20240124 2:16 PM] David, Lance (CTR) - FS, CO
        #Hi. The second one is correct. You will be computing BASA11 every cycle and then if BASA11 > 0 will compute BASA0and1 = 1
        #I would shorten the variable name BASA0and1 though, because FVS only acknowleges the frist 8 characters of variable names. It will not cause an error though.
        for k in range(0,len(BASA_VariableNameList),1):
            BASA_VariableName = BASA_VariableNameList[k]
            BASA_Species = BASA_SpeciesList[k]
            BASA_Size = BASA_SizeList[k]
            SpeciesSizeSpecficMetricsKcpFile.write("\n")
            SpeciesSizeSpecficMetricsKcpFile.write("if"+"\n")
            SpeciesSizeSpecficMetricsKcpFile.write(BASA_VariableName + " gt 0"+"\n")
            SpeciesSizeSpecficMetricsKcpFile.write("then"+"\n")
            SpeciesSizeSpecficMetricsKcpFile.write("Compute            0"+"\n")
            BASA0and1 = BASA_Species +"0A1"   #20240125, FVS will name the field as 0A1 if the A is a here, so it is important to use A for consistency
            SpeciesSizeSpecficMetricsKcpFile.write(BASA0and1 + " = 1"+"\n")
            SpeciesSizeSpecficMetricsKcpFile.write("end"+"\n")
            SpeciesSizeSpecficMetricsKcpFile.write("endif"+"\n")
            NewNameBASA0and1 = BASA_Species +"0AND1"+"_"+BASA_Size   #20240131: We will use this name to create a shape file in F3, but GDAL requires the metrci name maximum length is 10, so I change 0AND1 to 01
            CommmentLine = "* The FVS only allows maximum 8 charaters, so variable name of " + BASA0and1 + " below is actually " + NewNameBASA0and1   #This sentence 
            SpeciesSizeSpecficMetricsKcpFile.write(CommmentLine+"\n")
            SpeciesSizeSpecficMetricsKcpFile.write("if"+"\n")
            SpeciesSizeSpecficMetricsKcpFile.write(BASA_VariableName + " eq 0"+"\n")
            SpeciesSizeSpecficMetricsKcpFile.write("then"+"\n")
            SpeciesSizeSpecficMetricsKcpFile.write("Compute            0"+"\n")
            BASA0and1 = BASA_Species +"0A1"  #20240125, FVS will name the field as 0A1 if the A is a here, so it is important to use A for consistency
            SpeciesSizeSpecficMetricsKcpFile.write(BASA0and1 + " = 0"+"\n")
            SpeciesSizeSpecficMetricsKcpFile.write("end"+"\n")
            SpeciesSizeSpecficMetricsKcpFile.write("endif"+"\n")
        #202401424: This is to calculate the species 0 absence and 1 presence based on basal area value. The purpose is to impute the species confidence requested by Marla---end
        
        SpeciesSizeSpecficMetricsKcpFile.close()
        return SpeciesSizeSpecficMetricsKcp
    except Exception as e:
        print("Sorry, there is an error : ",e)
        exc_type, exc_value, exc_traceback = sys.exc_info()
        strTrace = traceback.format_exception(exc_type, exc_value, exc_traceback)
        ErrorLog = os.getcwd()+os.sep+"FvsError.txt"
        ErrorLogFile = open(ErrorLog, 'a')
        Content1 = "Dr. Shengli Huang detected an error at "+datetime.datetime.now().strftime("%Y-%m-%d %H:%M:%S") + "\n"
        #print(Content1)
        Content2 = "The error was for SpeciesSizeSpecficMetricsKcp with inputs of " + repr(SpeciesList)+","+repr(SizeList)+repr(MetricList) + " with the following details:\n"
        #print(Content2)
        Content3 = repr(strTrace) + "\n\n\n"  #see https://docs.python.org/2/library/traceback.html
        #print(Content3)
        ErrorLogFile.write(Content1)
        ErrorLogFile.write(Content2)
        ErrorLogFile.write(Content3)
        ErrorLogFile.close()
        return ErrorLog


def AutomaticUpdateF3MetricInformationExcel(F3MetricExcelXLSXfile,SpeciesSpecificAndSizeSpecificShortName,SpeciesSpecificAndSizeSpecificFullName):
    try:
        epoch_time = int(time.time())
        F3MetricExcelXLSXfileCopy = F3MetricExcelXLSXfile.replace(".xlsx","_epochtime"+str(epoch_time)+".xlsx")
        os.rename(F3MetricExcelXLSXfile,F3MetricExcelXLSXfileCopy)
        wb = openpyxl.load_workbook(F3MetricExcelXLSXfileCopy)  # Load the workbook
        ws = wb['F3MetricInfo'] #ws = wb.active is to Select the active worksheet, but here we select a specific worksheet. also 1) wb.sheetnames can get the worksheet names, 2) ws['A1'].value can get the cell value
        FieldLength = 0
        for cell in ws[1]:
            FieldLength = FieldLength + 1
            print("The field names are:", cell.value)
        ExistingMetricShortName = []
        rowid = 0
        for row in ws.iter_rows(min_row=2, values_only=True):  # Iterate through rows starting from the second row (i.e., ignore the first row which are field names
            rowid = rowid + 1
            ExistingMetricShortName.append(row[0])
        AddingMetricShortName = [x for x in SpeciesSpecificAndSizeSpecificFullName if x not in ExistingMetricShortName]
        print("We need add the following MetricShortName:",AddingMetricShortName)
        new_row = ["None"]*FieldLength 
        for k in AddingMetricShortName:
            if "0AND1_0_999" in k:
                Species = k.replace("0AND1_0_999","")
                new_row[0] = k
                new_row[1] = "Presence or absence of species " + Species + " with DBH between 0 to 999 inches"
                new_row[2] = new_row[1] + " referecned from FVS SPMCDBH()"
                new_row[3] = "Please check FIA_FVS_Species crosswalk for SpeciesScienticName"
                new_row[4] = "Please check FIA_FVS_Species crosswalk for USDAPlantCode"
                new_row[5] = "Self calculation in KCP based on basal area"
                new_row[6] = "0 or 1 imputed to range of 0-1"   #MetricUnit_US
                new_row[7] = "0 or 1 imputed to range of 0-1"   #MetricUnit_SI
                new_row[8] = "1"   #UStoSIconversion
                new_row[9] = "Shengli Huang (shengli.huang@usda.gov)"
                new_row[10] = "Shengli Huang (shengli.huang@usda.gov)"			
            else:
                new_row[0] = k
                new_row[1] = k.split("_")[0] + " of " + k.split("_")[1] + " with DBH between " + k.split("_")[2] + " and " + k.split("_")[3]
                new_row[2] = new_row[1] + " referecned from FVS SPMCDBH()"
                new_row[3] = "Please check FVS AlphaCode "+ k.split("_")[1] +" for SpeciesScienticName"
                new_row[4] = "Please check FIA_FVS_Species crosswalk for USDAPlantCode"
                new_row[5] = "SPMCDBH()"
                if k.split("_")[0] == "TPA":
                    new_row[6] = "Trees per acre"   #MetricUnit_US
                    new_row[7] = "trees per hectare"   #MetricUnit_SI
                    new_row[8] = "2.47105381"   #UStoSIconversion
                if k.split("_")[0] == "BASA":
                    new_row[6] = "square foot per acre"   #MetricUnit_US
                    new_row[7] = "square meters per hectare"   #MetricUnit_SI
                    new_row[8] = "0.229568"   #UStoSIconversion
                if k.split("_")[0] == "TCFVOL":
                    new_row[6] = "Cubic foot volume per acre"   #MetricUnit_US
                    new_row[7] = "Cubic meter volume per hectare"   #MetricUnit_SI
                    new_row[8] = "0.0699725"   #UStoSIconversion
                if k.split("_")[0] == "MBFVOL":
                    new_row[6] = "Cubic foot volume per acre"   #MetricUnit_US
                    new_row[7] = "Cubic meter volume per hectare"   #MetricUnit_SI
                    new_row[8] = "0.0699725"   #UStoSIconversion  
                if k.split("_")[0] == "QMD":
                    new_row[6] = "inch"   #MetricUnit_US
                    new_row[7] = "trees per hectare"   #MetricUnit_SI
                    new_row[8] = "2.54"   #UStoSIconversion
                if k.split("_")[0] == "AHGT":
                    new_row[6] = "feet"   #MetricUnit_US
                    new_row[7] = "meter"   #MetricUnit_SI
                    new_row[8] = "0.3048"   #UStoSIconversion
                if k.split("_")[0] == "CORPC":
                    new_row[6] = "Percentage"   #MetricUnit_US
                    new_row[7] = "Percentage"   #MetricUnit_SI
                    new_row[8] = "1"   #UStoSIconversion
                if k.split("_")[0] == "MISTLE":
                    new_row[6] = "rating"   #MetricUnit_US
                    new_row[7] = "rating"   #MetricUnit_SI
                    new_row[8] = "1"   #UStoSIconversion
                if k.split("_")[0] == "MCFVOL":
                    new_row[6] = "Cubic foot volume per acre"   #MetricUnit_US
                    new_row[7] = "Cubic meter volume per hectare"   #MetricUnit_SI
                    new_row[8] = "0.0699725"   #UStoSIconversion
                if k.split("_")[0] == "AVGDBHGROW":
                    new_row[6] = "inch"   #MetricUnit_US
                    new_row[7] = "trees per hectare"   #MetricUnit_SI
                    new_row[8] = "2.54"   #UStoSIconversion
                if k.split("_")[0] == "SDI1933":
                    new_row[6] = "number per acre"   #MetricUnit_US
                    new_row[7] = "number per acre"   #MetricUnit_SI
                    new_row[8] = "1"   #UStoSIconversion
                if k.split("_")[0] == "RELDEN1984":
                    new_row[6] = "number per acre"   #MetricUnit_US
                    new_row[7] = "number per acre"   #MetricUnit_SI
                    new_row[8] = "1"   #UStoSIconversion
                if k.split("_")[0] == "SDI1983":
                    new_row[6] = "number per acre"   #MetricUnit_US
                    new_row[7] = "number per acre"   #MetricUnit_SI
                    new_row[8] = "1"   #UStoSIconversion
                if k.split("_")[0] == "RELDEN2010":
                    new_row[6] = "number per acre"   #MetricUnit_US
                    new_row[7] = "number per acre"   #MetricUnit_SI
                    new_row[8] = "1"   #UStoSIconversion
                new_row[9] = "Shengli Huang (shengli.huang@usda.gov)"
                new_row[10] = "Shengli Huang (shengli.huang@usda.gov)"
            print("new_row=",new_row)
            ws.append(new_row)  #Adding New Records to an Excel Spreadsheet, we can use the append method
        wb.save(F3MetricExcelXLSXfile)
        wb.close()
        return "F3MetricExcelXLSXfile Updated"
    except Exception as e:
        print("Sorry, there is an error : ",e)
        exc_type, exc_value, exc_traceback = sys.exc_info()
        strTrace = traceback.format_exception(exc_type, exc_value, exc_traceback)
        ErrorLog = os.getcwd()+os.sep+"FvsError.txt"
        ErrorLogFile = open(ErrorLog, 'a')
        Content1 = "Dr. Shengli Huang detected an error at "+datetime.datetime.now().strftime("%Y-%m-%d %H:%M:%S") + "\n"
        #print(Content1)
        Content2 = "The error was for AutomaticUpdateF3MetricInformationExcel with inputs of " + repr(F3MetricExcelXLSXfile)+ repr(SpeciesSpecificAndSizeSpecificShortName)+ repr(SpeciesSpecificAndSizeSpecificFullName) + " with the following details:\n"
        #print(Content2)
        Content3 = repr(strTrace) + "\n\n\n"  #see https://docs.python.org/2/library/traceback.html
        #print(Content3)
        ErrorLogFile.write(Content1)
        ErrorLogFile.write(Content2)
        ErrorLogFile.write(Content3)
        ErrorLogFile.close()
        return ErrorLog
    

def SpeciesSpecificAndSizeSpecificNamePairs(SpeciesSizeSpecficMetricsKcpReturned):
    try:
        print("XXXXXXSpeciesSizeSpecficMetricsKcpReturned=",SpeciesSizeSpecficMetricsKcpReturned)
        SpeciesSpecificAndSizeSpecificShortName = []
        SpeciesSpecificAndSizeSpecificFullName = []
        if SpeciesSizeSpecficMetricsKcpReturned != "SpeciesSizeSpecficMetrics is not desired":  #This is just a sentence to judge if we really have a SpeciesSizeSpecficMetricsKcp. It corresponds to the [SpeciesSizeSpecficMetricsKcpReturned = "SpeciesSizeSpecficMetrics is not desired"] in main section 
            SpeciesSizeSpecficMetricsKcpReturnedFile = open(SpeciesSizeSpecficMetricsKcpReturned, 'r')
            Lines = SpeciesSizeSpecficMetricsKcpReturnedFile.readlines()
            SpeciesSizeSpecficMetricsKcpReturnedFile.close()
            Lines = [i.replace('\n','') for i in Lines]
            for eachline in Lines:
                #print("eachline=",eachline)
                if eachline.startswith("* The FVS only allows maximum 8 charaters"):
                    ShortName = eachline[eachline.find("so variable name of ")+len("so variable name of "):eachline.rfind("below is actually")].replace(" ", "")
                    FullName = eachline[eachline.find(" below is actually ")+len(" below is actually "):].replace(" ", "")
                    #print("ShortName,FullName=",ShortName,FullName)
                    SpeciesSpecificAndSizeSpecificShortName.append(ShortName)
                    SpeciesSpecificAndSizeSpecificFullName.append(FullName)
        return SpeciesSpecificAndSizeSpecificShortName,SpeciesSpecificAndSizeSpecificFullName
    except Exception as e:
        print("Sorry, there is an error : ",e)
        exc_type, exc_value, exc_traceback = sys.exc_info()
        strTrace = traceback.format_exception(exc_type, exc_value, exc_traceback)
        ErrorLog = os.getcwd()+os.sep+"FvsError.txt"
        ErrorLogFile = open(ErrorLog, 'a')
        Content1 = "Dr. Shengli Huang detected an error at "+datetime.datetime.now().strftime("%Y-%m-%d %H:%M:%S") + "\n"
        #print(Content1)
        Content2 = "The error was for SpeciesSpecificAndSizeSpecificNamePairs with inputs of " + repr(SpeciesSizeSpecficMetricsKcpReturned) + " with the following details:\n"
        #print(Content2)
        Content3 = repr(strTrace) + "\n\n\n"  #see https://docs.python.org/2/library/traceback.html
        #print(Content3)
        ErrorLogFile.write(Content1)
        ErrorLogFile.write(Content2)
        ErrorLogFile.write(Content3)
        ErrorLogFile.close()
        return ErrorLog
    

def ChangeFieldNameInFvsSqliteOutputForSpeciesSpecificAndSizeSpecific(FvsSqliteOutputDB, SpeciesSizeSpecficMetricsKcpReturned,SpeciesList):  #AddNewFieldToFvsSqliteOutput(OutputDB, "FIADB_PLOT")
    try:
        print("&&&&&&&&&&&&&&&&&SpeciesSizeSpecficMetricsKcpReturned=",SpeciesSizeSpecficMetricsKcpReturned)
        if SpeciesSizeSpecficMetricsKcpReturned != "SpeciesSizeSpecficMetrics is not desired":
            SpeciesSpecificAndSizeSpecificShortName,SpeciesSpecificAndSizeSpecificFullName = SpeciesSpecificAndSizeSpecificNamePairs(SpeciesSizeSpecficMetricsKcpReturned)
            print("SpeciesSpecificAndSizeSpecificShortName=",SpeciesSpecificAndSizeSpecificShortName,file=open(F3Log, 'a'))
            print("SpeciesSpecificAndSizeSpecificFullName=",SpeciesSpecificAndSizeSpecificFullName,file=open(F3Log, 'a'))
            IndividualConn = sqlite3.connect(FvsSqliteOutputDB)
            IndividualCursor = IndividualConn.cursor()
            IndividualCursor.execute("SELECT name FROM sqlite_master WHERE type='table' ORDER BY name;")  #List all the tables in the sqlite database
            IndividualDB_tables = IndividualCursor.fetchall()   #https://www.geeksforgeeks.org/how-to-list-tables-using-sqlite3-in-python/ and https://stackoverflow.com/questions/31986520/show-tables-in-sqlite-database-in-python
            IndividualDB_tables = [k[0] for k in IndividualDB_tables]  #IndividualDB_tables= [('FVS_Cases',), ('FVS_Summary2',), ('FVS_TreeList',)] so now IndividualDB_tables= ['FVS_Cases', 'FVS_Summary2', 'FVS_TreeList']
            for IndividualDB_table in IndividualDB_tables:
                TableQuery = 'SELECT * FROM ' + IndividualDB_table
                IndividualCursor.execute(TableQuery)   
                columns = [column[0] for column in IndividualCursor.description]

                #20240229 added: The "LT0and1_0_999" will be truncted as "LT0and1_" in FVS (max 8 characters, so I need to change it back---start
                if (("LT0AND1_".lower() in columns) or ("LT0AND1_" in columns)):  #Double check if 
                    TableFieldChangeQuery = "ALTER TABLE " + IndividualDB_table + " RENAME COLUMN " + "LT0AND1_" + " TO " + "LT0AND1_0_999"
                    IndividualCursor.execute(TableFieldChangeQuery)    
                #20240229 added: The "LT0and1_0_999" will be truncted as "LT0and1_" in FVS (max 8 characters, so I need to change it back---end

                #20240213: remove the field of SPECIES short name in the FVS result Sqlite DB---start
                #20240213: Lance David and I found the FVS resulting Table can add a new field like "PB", which create trouble when I tried to merge because the number of columns are different.
                #20240213: We found if the species is not in a variant's species list, then this happen (e.g., PB not in LS variant. We can check it in the OUT file [NATIONAL VOLUME ESTIMATOR LIBRARY EQUATION NUMBERS]) 
                #20240213: Therefore, I decide to remove the field of SPECIES (e.g., PB, BF, TA in central lakes variant). We never use SPECIES short name as the field, so we are safe to do it
                #https://www.sqlitetutor.com/drop-column/#google_vignette, and https://www.learningaboutelectronics.com/Articles/How-to-delete-a-column-of-a-MySQL-table-in-Python.php
                for species in SpeciesList:
                    if species in columns:
                        DropFieldSentence = "ALTER TABLE "+IndividualDB_table+" DROP "+species
                        print(FvsSqliteOutputDB," table ",IndividualDB_table," DropField ",species)
                        print(FvsSqliteOutputDB," table ",IndividualDB_table," DropField ",species,file=open(F3Log, 'a'))
                        IndividualCursor.execute(DropFieldSentence)
                #20240213: remove the field of SPECIES short name in the FVS result Sqlite DB---end
                
                #print(IndividualDB_table," has columns of ",columns)
                for k in range(0,len(SpeciesSpecificAndSizeSpecificShortName),1):
                    SpeciesSpecificAndSizeSpecificShortNameElement = SpeciesSpecificAndSizeSpecificShortName[k]
                    SpeciesSpecificAndSizeSpecificFullNameElement = SpeciesSpecificAndSizeSpecificFullName[k]
                    if SpeciesSpecificAndSizeSpecificShortNameElement in columns:
                        TableFieldChangeQuery = "ALTER TABLE " + IndividualDB_table + " RENAME COLUMN " + SpeciesSpecificAndSizeSpecificShortNameElement + " TO " + SpeciesSpecificAndSizeSpecificFullNameElement
                        #print("For ",FvsSqliteOutputDB, " we execute ", TableFieldChangeQuery,", i.e., the Field ",SpeciesSpecificAndSizeSpecificShortNameElement," is changed to ",SpeciesSpecificAndSizeSpecificFullNameElement)
                        IndividualCursor.execute(TableFieldChangeQuery)
            IndividualConn.commit()
            IndividualConn.close()
        return FvsSqliteOutputDB
    except Exception as e:
        print("Sorry, there is an error : ",e)
        exc_type, exc_value, exc_traceback = sys.exc_info()
        strTrace = traceback.format_exception(exc_type, exc_value, exc_traceback)
        ErrorLog = os.getcwd()+os.sep+"FvsError.txt"
        ErrorLogFile = open(ErrorLog, 'a')
        Content1 = "Dr. Shengli Huang detected an error at "+datetime.datetime.now().strftime("%Y-%m-%d %H:%M:%S") + "\n" 
        #print(Content1)
        Content2 = "The error was for ChangeFieldNameInFvsSqliteOutputForSpeciesSpecificAndSizeSpecific with inputs of " + repr(FvsSqliteOutputDB)+","+repr(SpeciesSizeSpecficMetricsKcpReturned)+","+repr(SpeciesList)+ " with the following details:\n"
        #print(Content2)
        Content3 = repr(strTrace) + "\n\n\n"  #see https://docs.python.org/2/library/traceback.html
        #print(Content3)
        ErrorLogFile.write(Content1)
        ErrorLogFile.write(Content2)
        ErrorLogFile.write(Content3)
        ErrorLogFile.close()
        return ErrorLog


def TIMEINTANDNUMCYCLE_function(INV_YEAR,FVS_Years,GrowthAndReportingInterval):  #check value and format please
    try:
        TimeInt = []
        FoundID = 0
        TimeIntSentence0 = "TimeInt                 " + str(GrowthAndReportingInterval)  #There are 17 spaces between after TimeInt to indicate the GrowthAndReportingInterval    
        TimeInt.append(TimeIntSentence0)
        for k in range(0,len(FVS_Years),1):
            if FVS_Years[k] > INV_YEAR:  # >= is not correct here. Note FVS always output the inventory year as the first
                FoundID = FoundID + 1
                if FoundID == 1:
                    YearDifference = FVS_Years[k] - INV_YEAR
                else:
                    YearDifference = FVS_Years[k] - FVS_Years[k-1]
                TimeIntSentence1 = "TimeInt       " + str(FoundID)  #There are 7 space after TimeInt
                TimeIntSentence2 = "        " + str(YearDifference) #There are 8 space before YearDifference
                TimeIntSentence = TimeIntSentence1 + TimeIntSentence2
                #if YearDifference != GrowthAndReportingInterval:  #if the YearDifference == GrowthAndReportingInterval, we will not write this line in Key file (I checked their key file and know it)
                TimeInt.append(TimeIntSentence)  
        TIMEINTANDNUMCYCLE = "\n".join(TimeInt) + "\n" + "NumCycle     "+str(FoundID+1)  #There are 5 space after NumCycle. 20230612: FoundID is changed to FoundID+1
        FVS_YearsText = [str(k) for k in FVS_Years]
        if INV_YEAR <= FVS_Years[0]:
            MasterKeyFileTemplateContentHead = "!!Expected FVS years are: " + ",".join(FVS_YearsText)+" and all years are modelled"
        if INV_YEAR > FVS_Years[0]:
            SkippedFVSYears = [str(k) for k in FVS_Years if k < INV_YEAR]
            MasterKeyFileTemplateContentHead = "!!Expected FVS years are: " + ",".join(FVS_YearsText)+" but the following years are skipped: " + ",".join(SkippedFVSYears) + ", because they are older than Inventory year of " + str(INV_YEAR)  
        return [TIMEINTANDNUMCYCLE, MasterKeyFileTemplateContentHead]
    except Exception as e:
        print("Sorry, there is an error : ",e)
        exc_type, exc_value, exc_traceback = sys.exc_info()
        strTrace = traceback.format_exception(exc_type, exc_value, exc_traceback)
        ErrorLog = os.getcwd()+os.sep+"FvsError.txt"
        ErrorLogFile = open(ErrorLog, 'a')
        Content1 = "Dr. Shengli Huang detected an error at "+datetime.datetime.now().strftime("%Y-%m-%d %H:%M:%S") + "\n"
        #print(Content1)
        Content2 = "The error was for TIMEINTANDNUMCYCLE_function with inputs of " + repr(INV_YEAR)+","+repr(FVS_Years) +","+repr(GrowthAndReportingInterval)+ " with the following details:\n"
        #print(Content2)
        Content3 = repr(strTrace) + "\n\n\n"  #see https://docs.python.org/2/library/traceback.html
        #print(Content3)
        ErrorLogFile.write(Content1)
        ErrorLogFile.write(Content2)
        ErrorLogFile.write(Content3)
        ErrorLogFile.close()
        return ErrorLog



def AddNewFieldToFvsSqliteOutput(FvsSqliteOutputDB, NewFieldNameList):  #AddNewFieldToFvsSqliteOutput(OutputDB, "FIADB_PLOT")
    try:
        FIADB_PLOT = FvsSqliteOutputDB.split(os.sep)[-1].split("_")[3][0:14]  #revised on 20240125 because the plotID is not 5 digits but total 14 digits (i.e., 2 UNIT + 4 state + 3 county + 5 plot)
        print("@@@@@@@@@@@@@@@@@@@@@@@@FIADB_PLOT = ", FIADB_PLOT, " for the SQLITEDB of ", FvsSqliteOutputDB,file=open(F3Log, 'a'))
        IndividualConn = sqlite3.connect(FvsSqliteOutputDB)
        IndividualCursor = IndividualConn.cursor()
        IndividualCursor.execute("SELECT name FROM sqlite_master WHERE type='table' ORDER BY name;")  #List all the tables in the sqlite database
        IndividualDB_tables = IndividualCursor.fetchall()   #https://www.geeksforgeeks.org/how-to-list-tables-using-sqlite3-in-python/ and https://stackoverflow.com/questions/31986520/show-tables-in-sqlite-database-in-python
        IndividualDB_tables = [k[0] for k in IndividualDB_tables]  #IndividualDB_tables= [('FVS_Cases',), ('FVS_Summary2',), ('FVS_TreeList',)] so now IndividualDB_tables= ['FVS_Cases', 'FVS_Summary2', 'FVS_TreeList']
        for IndividualDB_table in IndividualDB_tables:
            TableQuery = 'SELECT * FROM ' + IndividualDB_table
            IndividualCursor.execute(TableQuery)   #Here we can check if a table have "Year. If yes, only select the Year; if not, then copy the whole table?
            columns = [column[0] for column in IndividualCursor.description]
            for NewFieldName in NewFieldNameList:
                if NewFieldName not in columns:
                    if NewFieldName == "FIADB_PLOT":
                        column_type = 'INTEGER'
                        FieldValue = str(FIADB_PLOT)
                    elif NewFieldName == "Source":
                        column_type = 'TEXT'
                        FieldValue = '"' + FvsSqliteOutputDB.split(os.sep)[-1].replace(".db","") + '"'
                    elif NewFieldName == "Category":
                        column_type = 'TEXT'
                        FieldValue = '"' + FvsSqliteOutputDB.split(os.sep)[-1].split("_")[-1].replace(".db","") + '"'                        
                    else:
                        column_type = 'TEXT'
                        FieldValue = "PlaceHolder"
                    AddFieldCommandLine = 'ALTER TABLE ' + IndividualDB_table + ' ADD COLUMN ' + NewFieldName + ' ' + column_type
                    #print("AddFieldCommandLine: ",AddFieldCommandLine,file=open(F3Log, 'a'))
                    IndividualCursor.execute(AddFieldCommandLine)   #'''ALTER TABLE table_name ADD COLUMN column_name column_type'''
                    AssignValueToFieldCommandLine = "UPDATE " + IndividualDB_table + " SET " + NewFieldName + " = " + FieldValue
                    #print("AssignValueToFieldCommandLine: ",AssignValueToFieldCommandLine,file=open(F3Log, 'a'))
                    IndividualCursor.execute(AssignValueToFieldCommandLine) 
        IndividualConn.commit()
        IndividualConn.close()
        return FvsSqliteOutputDB
    except Exception as e:
        print("Sorry, there is an error : ",e)
        exc_type, exc_value, exc_traceback = sys.exc_info()
        strTrace = traceback.format_exception(exc_type, exc_value, exc_traceback)
        ErrorLog = os.getcwd()+os.sep+"FvsError.txt"
        ErrorLogFile = open(ErrorLog, 'a')
        Content1 = "Dr. Shengli Huang detected an error at "+datetime.datetime.now().strftime("%Y-%m-%d %H:%M:%S") + "\n" 
        #print(Content1)
        Content2 = "The error was for AddNewFieldToFvsSqliteOutput with inputs of " + repr(FvsSqliteOutputDB)+","+repr(NewFieldNameList)+ " with the following details:\n"
        #print(Content2)
        Content3 = repr(strTrace) + "\n\n\n"  #see https://docs.python.org/2/library/traceback.html
        #print(Content3)
        ErrorLogFile.write(Content1)
        ErrorLogFile.write(Content2)
        ErrorLogFile.write(Content3)
        ErrorLogFile.close()
        return ErrorLog

    
def CreateFIADBUniquePlotIDfromFvsReadyStandID(STAND_ID):
    try:
        #note in FVS_ready database, Stand_ID = concatenate((PLOT.STATECD(4) + PLOT.INVYR(4) + PLOT.CYCLE(2) + PLOT.SUBCYCLE(2) + PLOT.UNITCD(2) + PLOT.COUNTYCD(3) + PLOT.PLOT(5)).
        StateCode = STAND_ID[0:4]  #The first 4 digits in FIA_FVSready FVS_STANDINIT_PLOT table is state code. Four digits
        #UnitCode is usually like 03, 05, 08 etc. Below I change it to 13, 15, and 18 etc to make sure the FIADB_PLOT will always have the same number of digits for processing
        UnitCode = str(int(STAND_ID[12:14])+10)  #Survey unit code. Forest Inventory and Analysis survey unit identification number. Survey units are usually groups of counties within each State. Two digits
        CountyCode = STAND_ID[14:17]  ##The last 8 digits to 5 digits in FIA_FVSready FVS_STANDINIT_PLOT table is county code. Three digits
        PlotCode = STAND_ID[-5:]   #The last 5 digits in FIA_FVSready FVS_STANDINIT_PLOT table is plot code. Five digits
        FIADB_PLOT = UnitCode + StateCode + CountyCode +  PlotCode   #Since 20231229, I decided to create unique plot id. The original STABD_ID=0001197204000300190045, then FIADB_PLOT= 13000100190045 
        return FIADB_PLOT
    except Exception as e:
        print("Sorry, there is an error : ",e)
        exc_type, exc_value, exc_traceback = sys.exc_info()
        strTrace = traceback.format_exception(exc_type, exc_value, exc_traceback)
        ErrorLog = os.getcwd()+os.sep+"FvsError.txt"
        ErrorLogFile = open(ErrorLog, 'a')
        Content1 = "Dr. Shengli Huang detected an error at "+datetime.datetime.now().strftime("%Y-%m-%d %H:%M:%S") + "\n" 
        #print(Content1)
        Content2 = "The error was for CreateFIADBUniquePlotIDfromFvsReadyStandID with inputs of " + repr(STAND_ID)+ " with the following details:\n"
        #print(Content2)
        Content3 = repr(strTrace) + "\n\n\n"  #see https://docs.python.org/2/library/traceback.html
        #print(Content3)
        ErrorLogFile.write(Content1)
        ErrorLogFile.write(Content2)
        ErrorLogFile.write(Content3)
        ErrorLogFile.close()
        return ErrorLog


def FVSKeyBatAndRun(mylock,Tile,Category,Management,ManagementHeader,SelectFIAplotMeasurement,MasterKeyFileTemplate,FVSBatFileTemplate,FVSINPUTDBs,FVS_Years,OutputKcpList,EventKcpList,SpeciesSizeSpecficMetricsKcpReturned,SpeciesList):
    try:
        FVSOUTPUTDB = FVSprocessingpath + os.sep + Management + os.sep + Tile + os.sep + ManagementHeader + "_"+ Management + "_"+SelectFIAplotMeasurement+"_"+Category+".db"  #the first five digits of SelectFIAplotMeasurement is plot ID
        if not os.path.exists(FVSOUTPUTDB):
            
            #if not SelectFIAplotMeasurement.startswith("999"):  #82330 is a good plot to check
            #    return
            
            print("SelectFIAplotMeasurement from FVSKeyBatAndRun is ",SelectFIAplotMeasurement,file=open(F3Log, 'a'))
            print("SelectFIAplotMeasurement from FVSKeyBatAndRun is ",SelectFIAplotMeasurement)
            MasterKeyFileTemplateFile = open(MasterKeyFileTemplate,mode='r')  #https://www.tutorialspoint.com/how-to-read-a-text-file-in-python
            MasterKeyFileTemplateContentMain = MasterKeyFileTemplateFile.read()
            MasterKeyFileTemplateFile.close()

            FVSBatFileTemplateFile = open(FVSBatFileTemplate,mode='r')  #https://www.tutorialspoint.com/how-to-read-a-text-file-in-python
            FVSBatFileTemplateContentMain = FVSBatFileTemplateFile.read()
            FVSBatFileTemplateFile.close()
            for FVSINPUTDB in FVSINPUTDBs:
                conn = sqlite3.connect(FVSINPUTDB)  #https://www.sqlitetutorial.net/sqlite-python/
                cur= conn.cursor()
                cur.execute("SELECT * FROM FVS_STANDINIT_PLOT")  #Ask Marcus to give the table name consistently
                columns = [column[0] for column in cur.description]
                #print("columns=",columns)
                STAND_IDIndex = columns.index("STAND_ID")
                #print("STAND_IDIndex=",STAND_IDIndex)
                STAND_CNIndex = columns.index("STAND_CN")
                INV_YEARIndex = columns.index("INV_YEAR")
                INV_MONTHIndex = columns.index("INV_MONTH")
                INV_DAYIndex = columns.index("INV_DAY")
                VARIANTIndex = columns.index("VARIANT")
                LATITUDEIndex = columns.index("LATITUDE")
                LONGITUDEIndex = columns.index("LONGITUDE")
                ELEVFTIndex = columns.index("ELEVFT")
                rows = cur.fetchall()

                rowid = 0
                for row in rows:
                    rowid = rowid + 1
                    #print("rowid=",rowid)
                    TITLE = "HPM Michigan project" #(max length of 72)
                    STAND_ID = str(row[STAND_IDIndex])
                    STAND_CN = str(row[STAND_CNIndex])
                    INV_YEAR = int(row[INV_YEARIndex])
                    INV_MONTH = row[INV_MONTHIndex]
                    if INV_MONTH is None:   #This is consistent with previous work of HSL and month assignment
                        INV_MONTH = 7
                    else:
                        INV_MONTH = int(INV_MONTH)
                    INV_DAY = row[INV_DAYIndex]
                    if INV_DAY is None:     #This is consistent with previous work of HSL and day assignment
                        INV_DAY = 15
                    else:
                        INV_DAY = int(INV_DAY)
                    VARIANT = str(row[VARIANTIndex])
                    LATUTUDE = str(row[LATITUDEIndex])
                    LONGITUDE = str(row[LONGITUDEIndex])
                    ELEVFT = str(row[ELEVFTIndex])
                    STAND_IDANDTITLE = STAND_ID + " " * (27 - len(STAND_ID)) + TITLE  #From system-generated key file, I found the TITLE always starts at 28th position
                    #FIADB_PLOT = STAND_ID[-5:]  #STAND_ID[-5:] is to get the PlotFIADB from stand such as 0006200105010500350682 whose last 5 digits are FiadbPlot
                    FIADB_PLOT = CreateFIADBUniquePlotIDfromFvsReadyStandID(STAND_ID)   #20240102 added to replace th sentence above
                    FiaMeasurementTime = FIADB_PLOT + str(INV_MONTH).zfill(2) + str(INV_YEAR)  ##see zfill at https://stackoverflow.com/questions/134934/display-number-with-leading-zeros
                    if FiaMeasurementTime != SelectFIAplotMeasurement:
                        continue

                    DESCRIPTION = "Plot("+FIADB_PLOT+")-MeasuredYear("+str(INV_YEAR)+")-MeasuredMonth("+str(INV_MONTH)+")-MeasuredDay("+str(INV_DAY)+")-FVS VARIANT("+VARIANT+")-Fuzzy Latitude("+LATUTUDE+")-Fuzzy Longitude("+LONGITUDE+")-Fuzzy FT("+ELEVFT+")"
                    print("DESCRIPTION=",DESCRIPTION)
                    GrowthAndReportingInterval = 10
                    TIMEINTANDNUMCYCLEHeader = TIMEINTANDNUMCYCLE_function(INV_YEAR,FVS_Years,GrowthAndReportingInterval)
                    F3Paper = r'Huang, Shengli, Carlos Ramirez, Marcus McElhaney, and Kirk Evans. "F3: Simulating spatiotemporal forest change from field inventory, remote sensing, growth modeling, and management actions." Forest Ecology and Management 415 (2018): 26-37'
                    TIME = "Created by Dr. Shengli Huang at "+datetime.datetime.now().strftime("%Y-%m-%d %H:%M:%S") + " for F3 modeling (see " + F3Paper +")"

                    KCP_element = []
                    KCP_element.append("!!There are "+ str(len(EventKcpList)) +" EventKcp list---------------------------------------start")
                    for n in range(0,len(EventKcpList),1):
                        Kcp1 = "OPEN        " + str(60+n) + "\n"
                        Kcp2 = EventKcpList[n] + "\n"
                        Kcp3 = "ADDFILE     " + str(60+n) + "\n"
                        Kcp4 = "CLOSE       " + str(60+n)
                        KCPLine = Kcp1 + Kcp2 + Kcp3 + Kcp4
                        KCP_element.append(KCPLine)
                    KCP_element.append("!!There are "+ str(len(EventKcpList)) +" EventKcp list--------------------------------------end")
                    KCP_element.append("!!There are "+ str(len(OutputKcpList)) +" OutputKcp list-------------------------------------start")
                    for m in range(0,len(OutputKcpList),1):
                        Kcp1 = "OPEN        " + str(60+len(EventKcpList)+m) + "\n"
                        Kcp2 = OutputKcpList[m] + "\n"
                        Kcp3 = "ADDFILE     " + str(60+len(EventKcpList)+m) + "\n"
                        Kcp4 = "CLOSE       " + str(60+len(EventKcpList)+m)  
                        KCPLine = Kcp1 + Kcp2 + Kcp3 + Kcp4
                        KCP_element.append(KCPLine)
                    KCP_element.append("!!There are "+ str(len(OutputKcpList)) +" OutputKcp list-------------------------------------end")
                    
                    MORTALITYEVENT = "!!No special mortality event"

                    
                    KCPCONTENT = "\n".join(KCP_element)
                    OutputKey = FVSOUTPUTDB.replace(".db",".key")
                    OutputKeyFile = open(OutputKey,mode='w')  #https://www.tutorialspoint.com/how-to-read-a-text-file-in-python

                    Header = str(TIMEINTANDNUMCYCLEHeader[1]) + "\n"
                    OutputKeyFile.write(Header)
                    MasterKeyFileTemplateContent = MasterKeyFileTemplateContentMain
                    MasterKeyFileTemplateContent = MasterKeyFileTemplateContent.replace("$TITLE$",TITLE)
                    MasterKeyFileTemplateContent = MasterKeyFileTemplateContent.replace("$STAND_IDANDTITLE$",STAND_IDANDTITLE)
                    MasterKeyFileTemplateContent = MasterKeyFileTemplateContent.replace("$FVSINPUTDB$",FVSINPUTDB)
                    MasterKeyFileTemplateContent = MasterKeyFileTemplateContent.replace("$DESCRIPTION$",DESCRIPTION)
                    MasterKeyFileTemplateContent = MasterKeyFileTemplateContent.replace("$TIME$",TIME)
                    MasterKeyFileTemplateContent = MasterKeyFileTemplateContent.replace("$STAND_ID$",STAND_ID)
                    MasterKeyFileTemplateContent = MasterKeyFileTemplateContent.replace("$STAND_CN$",STAND_CN)
                    MasterKeyFileTemplateContent = MasterKeyFileTemplateContent.replace("$INV_YEAR$",str(INV_YEAR))
                    MasterKeyFileTemplateContent = MasterKeyFileTemplateContent.replace("$FVSOUTPUTDB$",FVSOUTPUTDB)
                    MasterKeyFileTemplateContent = MasterKeyFileTemplateContent.replace("$KCPCONTENT$",KCPCONTENT)
                    MasterKeyFileTemplateContent = MasterKeyFileTemplateContent.replace("$MORTALITYEVENT$",MORTALITYEVENT)
                    TIMEINTANDNUMCYCLE = str(TIMEINTANDNUMCYCLEHeader[0])
                    MasterKeyFileTemplateContent = MasterKeyFileTemplateContent.replace("$TIMEINTANDNUMCYCLE$",TIMEINTANDNUMCYCLE)
                    OutputKeyFile.write(MasterKeyFileTemplateContent)
                    OutputKeyFile.close()
                    print("Key file just created: ", OutputKey)

                    
                    FVSKEYFILEWithoutExtension = OutputKey.replace(".key","")
                    #print("FVSKEYFILEWithoutExtension=",FVSKEYFILEWithoutExtension)
                    OutputBat = OutputKey.replace(".key",".bat")
                    OutputBatFile = open(OutputBat,mode='w')  #https://www.tutorialspoint.com/how-to-read-a-text-file-in-python
                    FVSBatFileTemplateContent = FVSBatFileTemplateContentMain
                    FVSBatFileTemplateContent = FVSBatFileTemplateContent.replace("$FVSKEYFILEWithoutExtension$",FVSKEYFILEWithoutExtension)
                    FVSvariantExe = "FVS" + VARIANT + ".exe"
                    FVSBatFileTemplateContent = FVSBatFileTemplateContent.replace("$FVSvariantExe$",FVSvariantExe)
                    OutputBatFile.write(FVSBatFileTemplateContent)
                    OutputBatFile.close()
                    print("Bat file just created:", OutputBat)

                    time.sleep(30)
                    ThisBatFile = OutputBat
                    BatFileName = ThisBatFile.split(os.sep)[-1]
                    BatFilePath = ThisBatFile.replace(os.sep+ThisBatFile.split("\\")[-1],"")
                    #print("BatFilePath=",BatFilePath)
                    #print("BatFileName=",BatFileName)
                    OutFileOfBatchRun = ThisBatFile.replace(".bat", ".out")
                    p = Popen(BatFileName, cwd=BatFilePath,shell=True)   #Popen is using python subprocess module, but we can test os.system(ThisBatFile) in the future
                    stdout, stderr = p.communicate()
                    print("Successful in running FVS *.BAT file of: ", ThisBatFile)

                    #20230503 added. This section is to add a new field called FIADB_PLOT in the FVS output---start
                    OutputDB = OutputKey.replace(".key",".db")
                    NewFieldNameList = ["FIADB_PLOT","Source","Category"]
                    OutputDB111 = AddNewFieldToFvsSqliteOutput(OutputDB, NewFieldNameList)
                    #20230503 added. This section is to add a new field called FIADB_PLOT in the FVS output---end

                    #20240114 added. This section is to ChangeFieldNameInFvsSqliteOutputForSpeciesSpecificAndSizeSpecific because the FVS short variable name is not informative---start
                    #print("*****************OutputDB111 done*****************************:",OutputDB111)
                    OutputDB = ChangeFieldNameInFvsSqliteOutputForSpeciesSpecificAndSizeSpecific(OutputDB111, SpeciesSizeSpecficMetricsKcpReturned,SpeciesList)
                    #20240114 added. This section is to ChangeFieldNameInFvsSqliteOutputForSpeciesSpecificAndSizeSpecific because the FVS short variable name is not informative---start
                
    except Exception as e:
        print("Sorry, there is an error : ",e)
        exc_type, exc_value, exc_traceback = sys.exc_info()
        strTrace = traceback.format_exception(exc_type, exc_value, exc_traceback)
        ErrorLog = os.getcwd()+os.sep+"FvsError.txt"
        ErrorLogFile = open(ErrorLog, 'a')
        Content1 = "Dr. Shengli Huang detected an error at "+datetime.datetime.now().strftime("%Y-%m-%d %H:%M:%S") + "\n" 
        #print(Content1)
        Content2 = "The error was for FVSKeyBatAndRun with inputs of " + repr(Tile)+","+ repr(SelectFIAplotMeasurement)+","+repr(MasterKeyFileTemplate)+","+repr(FVSBatFileTemplate)+","+repr(FVSINPUTDBs)+","+repr(SpeciesSizeSpecficMetricsKcpReturned)+ " with the following details:\n"
        #print(Content2)
        Content3 = repr(strTrace) + "\n\n\n"  #see https://docs.python.org/2/library/traceback.html
        #print(Content3)
        ErrorLogFile.write(Content1)
        ErrorLogFile.write(Content2)
        ErrorLogFile.write(Content3)
        ErrorLogFile.close()
        return ErrorLog



def FVSKeyBatAndRunDisturbedFVS(mylock,Tile,Category,Management,ManagementHeader,SelectFIAplotMeasurementDisturbedFVS,MasterKeyFileTemplate,FVSBatFileTemplate,FVSINPUTDBs,FVS_Years,OutputKcpList,EventKcpList,SpeciesSizeSpecficMetricsKcpReturned,SpeciesList):
    try:
        DisturbedYear = [k.split("/")[0] for k in SelectFIAplotMeasurementDisturbedFVS.split(",")[1:]]
        DisturbedMortalityPercentage = [k.split("/")[1] for k in SelectFIAplotMeasurementDisturbedFVS.split(",")[1:]]
        SelectFIAplotMeasurement = SelectFIAplotMeasurementDisturbedFVS.split(",")[0]
        print("DisturbedYear=",DisturbedYear)
        print("DisturbedMortalityPercentage=",DisturbedMortalityPercentage)
        print("SelectFIAplotMeasurement=",SelectFIAplotMeasurement)
        
        FVSOUTPUTDB = FVSprocessingpath + os.sep + Management + os.sep + Tile + os.sep + ManagementHeader + "_"+ Management + "_"+SelectFIAplotMeasurement+"_"+Category+".db"  #the first five digits of SelectFIAplotMeasurement is plot ID
        if not os.path.exists(FVSOUTPUTDB):

            print("SelectFIAplotMeasurement from FVSKeyBatAndRunDisturbedFVS is ",SelectFIAplotMeasurement,file=open(F3Log, 'a'))
            MasterKeyFileTemplateFile = open(MasterKeyFileTemplate,mode='r')  #https://www.tutorialspoint.com/how-to-read-a-text-file-in-python
            MasterKeyFileTemplateContentMain = MasterKeyFileTemplateFile.read()
            MasterKeyFileTemplateFile.close()

            FVSBatFileTemplateFile = open(FVSBatFileTemplate,mode='r')  #https://www.tutorialspoint.com/how-to-read-a-text-file-in-python
            FVSBatFileTemplateContentMain = FVSBatFileTemplateFile.read()
            FVSBatFileTemplateFile.close()
            for FVSINPUTDB in FVSINPUTDBs:
                conn = sqlite3.connect(FVSINPUTDB)  #https://www.sqlitetutorial.net/sqlite-python/
                cur= conn.cursor()
                cur.execute("SELECT * FROM FVS_STANDINIT_PLOT")  #Ask Marcus to give the table name consistently
                columns = [column[0] for column in cur.description]
                #print(columns)
                STAND_IDIndex = columns.index("STAND_ID")
                #print("STAND_IDIndex=",STAND_IDIndex)
                STAND_CNIndex = columns.index("STAND_CN")
                INV_YEARIndex = columns.index("INV_YEAR")
                INV_MONTHIndex = columns.index("INV_MONTH")
                INV_DAYIndex = columns.index("INV_DAY")
                VARIANTIndex = columns.index("VARIANT")
                LATITUDEIndex = columns.index("LATITUDE")
                LONGITUDEIndex = columns.index("LONGITUDE")
                ELEVFTIndex = columns.index("ELEVFT")
                rows = cur.fetchall()

                rowid = 0
                for row in rows:
                    rowid = rowid + 1
                    
                    TITLE = "ACCEL project" #(max length of 72)
                    STAND_ID = str(row[STAND_IDIndex])
                    STAND_CN = str(row[STAND_CNIndex])
                    INV_YEAR = int(row[INV_YEARIndex])
                    INV_MONTH = row[INV_MONTHIndex]
                    if INV_MONTH is None:   #This is consistent with previous work of HSL and month assignment
                        INV_MONTH = 7
                    else:
                        INV_MONTH = int(INV_MONTH)
                    INV_DAY = row[INV_DAYIndex]
                    if INV_DAY is None:     #This is consistent with previous work of HSL and day assignment
                        INV_DAY = 15
                    else:
                        INV_DAY = int(INV_DAY)
                    VARIANT = str(row[VARIANTIndex])
                    LATUTUDE = str(row[LATITUDEIndex])
                    LONGITUDE = str(row[LONGITUDEIndex])
                    ELEVFT = str(row[ELEVFTIndex])
                    STAND_IDANDTITLE = STAND_ID + " " * (27 - len(STAND_ID)) + TITLE  #From system-generated key file, I found the TITLE always starts at 28th position
                    #FIADB_PLOT = STAND_ID[-5:]  #STAND_ID[-5:] is to get the PlotFIADB from stand such as 0006200105010500350682 whose last 5 digits are FiadbPlot
                    FIADB_PLOT = CreateFIADBUniquePlotIDfromFvsReadyStandID(STAND_ID)   #20240102 added to replace th sentence above
                    
                    FiaMeasurementTime = FIADB_PLOT + str(INV_MONTH).zfill(2) + str(INV_YEAR)  ##see zfill at https://stackoverflow.com/questions/134934/display-number-with-leading-zeros
                    if FiaMeasurementTime != SelectFIAplotMeasurement:
                        continue

                    DESCRIPTION = "Plot("+FIADB_PLOT+")-MeasuredYear("+str(INV_YEAR)+")-MeasuredMonth("+str(INV_MONTH)+")-MeasuredDay("+str(INV_DAY)+")-FVS VARIANT("+VARIANT+")-Fuzzy Latitude("+LATUTUDE+")-Fuzzy Longitude("+LONGITUDE+")-Fuzzy FT("+ELEVFT+")"
                    GrowthAndReportingInterval = 10
                    TIMEINTANDNUMCYCLEHeader = TIMEINTANDNUMCYCLE_function(INV_YEAR,FVS_Years,GrowthAndReportingInterval)
                    F3Paper = r'Huang, Shengli, Carlos Ramirez, Marcus McElhaney, and Kirk Evans. "F3: Simulating spatiotemporal forest change from field inventory, remote sensing, growth modeling, and management actions." Forest Ecology and Management 415 (2018): 26-37'
                    TIME = "Created by Dr. Shengli Huang at "+datetime.datetime.now().strftime("%Y-%m-%d %H:%M:%S") + " for F3 modeling (see " + F3Paper +")"

                    KCP_element = []
                    KCP_element.append("!!There are "+ str(len(EventKcpList)) +" EventKcp list---------------------------------------start")
                    for n in range(0,len(EventKcpList),1):
                        Kcp1 = "OPEN        " + str(60+n) + "\n"
                        Kcp2 = EventKcpList[n] + "\n"
                        Kcp3 = "ADDFILE     " + str(60+n) + "\n"
                        Kcp4 = "CLOSE       " + str(60+n)
                        KCPLine = Kcp1 + Kcp2 + Kcp3 + Kcp4
                        KCP_element.append(KCPLine)
                    KCP_element.append("!!There are "+ str(len(EventKcpList)) +" EventKcp list--------------------------------------end")
                    KCP_element.append("!!There are "+ str(len(OutputKcpList)) +" OutputKcp list-------------------------------------start")
                    for m in range(0,len(OutputKcpList),1):
                        Kcp1 = "OPEN        " + str(60+len(EventKcpList)+m) + "\n"
                        Kcp2 = OutputKcpList[m] + "\n"
                        Kcp3 = "ADDFILE     " + str(60+len(EventKcpList)+m) + "\n"
                        Kcp4 = "CLOSE       " + str(60+len(EventKcpList)+m)  
                        KCPLine = Kcp1 + Kcp2 + Kcp3 + Kcp4
                        KCP_element.append(KCPLine)
                    KCP_element.append("!!There are "+ str(len(OutputKcpList)) +" OutputKcp list-------------------------------------end")
                    
                    MortalityLine = []
                    for k in range(0,len(DisturbedYear),1):
                        MortalityContent0 = "!Exten:base Title: Adjust mortality rates of " + str(DisturbedMortalityPercentage[k]) + " caused by the disturbance occurred in " + str(DisturbedYear[k])+ ". In the future we can make it more complicated and add regeneration"
                        MortalityLine.append(MortalityContent0)
                        MortalityContent1 = "FixMort         " + str(DisturbedYear[k]) +"       All        " + str(DisturbedMortalityPercentage[k]) + "       0.0     999.0         0         0"
                        MortalityLine.append(MortalityContent1)
                    MORTALITYEVENT = "\n".join(MortalityLine)

                    KCPCONTENT = "\n".join(KCP_element)
                    OutputKey = FVSOUTPUTDB.replace(".db",".key")
                    OutputKeyFile = open(OutputKey,mode='w')  #https://www.tutorialspoint.com/how-to-read-a-text-file-in-python

                    Header = str(TIMEINTANDNUMCYCLEHeader[1]) + "\n"
                    OutputKeyFile.write(Header)
                    MasterKeyFileTemplateContent = MasterKeyFileTemplateContentMain
                    MasterKeyFileTemplateContent = MasterKeyFileTemplateContent.replace("$TITLE$",TITLE)
                    MasterKeyFileTemplateContent = MasterKeyFileTemplateContent.replace("$STAND_IDANDTITLE$",STAND_IDANDTITLE)
                    MasterKeyFileTemplateContent = MasterKeyFileTemplateContent.replace("$FVSINPUTDB$",FVSINPUTDB)
                    MasterKeyFileTemplateContent = MasterKeyFileTemplateContent.replace("$DESCRIPTION$",DESCRIPTION)
                    MasterKeyFileTemplateContent = MasterKeyFileTemplateContent.replace("$TIME$",TIME)
                    MasterKeyFileTemplateContent = MasterKeyFileTemplateContent.replace("$STAND_ID$",STAND_ID)
                    MasterKeyFileTemplateContent = MasterKeyFileTemplateContent.replace("$STAND_CN$",STAND_CN)
                    MasterKeyFileTemplateContent = MasterKeyFileTemplateContent.replace("$INV_YEAR$",str(INV_YEAR))
                    MasterKeyFileTemplateContent = MasterKeyFileTemplateContent.replace("$FVSOUTPUTDB$",FVSOUTPUTDB)
                    MasterKeyFileTemplateContent = MasterKeyFileTemplateContent.replace("$KCPCONTENT$",KCPCONTENT)
                    MasterKeyFileTemplateContent = MasterKeyFileTemplateContent.replace("$MORTALITYEVENT$",MORTALITYEVENT)
                    TIMEINTANDNUMCYCLE = str(TIMEINTANDNUMCYCLEHeader[0])
                    MasterKeyFileTemplateContent = MasterKeyFileTemplateContent.replace("$TIMEINTANDNUMCYCLE$",TIMEINTANDNUMCYCLE)
                    OutputKeyFile.write(MasterKeyFileTemplateContent)
                    OutputKeyFile.close()
                    print("Key file just created: ", OutputKey)

                    
                    FVSKEYFILEWithoutExtension = OutputKey.replace(".key","")
                    #print("FVSKEYFILEWithoutExtension=",FVSKEYFILEWithoutExtension)
                    OutputBat = OutputKey.replace(".key",".bat")
                    OutputBatFile = open(OutputBat,mode='w')  #https://www.tutorialspoint.com/how-to-read-a-text-file-in-python
                    FVSBatFileTemplateContent = FVSBatFileTemplateContentMain
                    FVSBatFileTemplateContent = FVSBatFileTemplateContent.replace("$FVSKEYFILEWithoutExtension$",FVSKEYFILEWithoutExtension)
                    FVSvariantExe = "FVS" + VARIANT + ".exe"
                    FVSBatFileTemplateContent = FVSBatFileTemplateContent.replace("$FVSvariantExe$",FVSvariantExe)
                    OutputBatFile.write(FVSBatFileTemplateContent)
                    OutputBatFile.close()
                    print("Bat file just created:", OutputBat)

                    time.sleep(30)
                    ThisBatFile = OutputBat
                    BatFileName = ThisBatFile.split(os.sep)[-1]
                    BatFilePath = ThisBatFile.replace(os.sep+ThisBatFile.split("\\")[-1],"")
                    #print("BatFilePath=",BatFilePath)
                    #print("BatFileName=",BatFileName)
                    OutFileOfBatchRun = ThisBatFile.replace(".bat", ".out")
                    p = Popen(BatFileName, cwd=BatFilePath,shell=True)   #Popen is using python subprocess module, but we can test os.system(ThisBatFile) in the future
                    stdout, stderr = p.communicate()
                    print("Successful in running FVS *.BAT file of: ", ThisBatFile)

                    #20230503 added. This section is to add a new field called FIADB_PLOT in the FVS output---start
                    OutputDB = OutputKey.replace(".key",".db")
                    NewFieldNameList = ["FIADB_PLOT","Source","Category"]
                    OutputDB111 = AddNewFieldToFvsSqliteOutput(OutputDB, NewFieldNameList)
                    #20230503 added. This section is to add a new field called FIADB_PLOT in the FVS output---end

                    #20240114 added. This section is to ChangeFieldNameInFvsSqliteOutputForSpeciesSpecificAndSizeSpecific because the FVS short variable name is not informative---start
                    print("$$$$$$$$$$$$$$$$$$$$$$$OutputDB111 done$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$")
                    OutputDB = ChangeFieldNameInFvsSqliteOutputForSpeciesSpecificAndSizeSpecific(OutputDB111, SpeciesSizeSpecficMetricsKcpReturned,SpeciesList)
                    #20240114 added. This section is to ChangeFieldNameInFvsSqliteOutputForSpeciesSpecificAndSizeSpecific because the FVS short variable name is not informative---start

    except Exception as e:
        print("Sorry, there is an error : ",e)
        exc_type, exc_value, exc_traceback = sys.exc_info()
        strTrace = traceback.format_exception(exc_type, exc_value, exc_traceback)
        ErrorLog = os.getcwd()+os.sep+"FvsError.txt"
        ErrorLogFile = open(ErrorLog, 'a')
        Content1 = "Dr. Shengli Huang detected an error at "+datetime.datetime.now().strftime("%Y-%m-%d %H:%M:%S") + "\n" 
        #print(Content1)
        Content2 = "The error was for FVSKeyBatAndRunDisturbedFVS with inputs of " + repr(Tile)+"," + repr(SelectFIAplotMeasurement)+","+repr(MasterKeyFileTemplate)+","+repr(FVSBatFileTemplate)+","+repr(FVSINPUTDBs)+","+repr(SpeciesSizeSpecficMetricsKcpReturned)+ " with the following details:\n"
        #print(Content2)
        Content3 = repr(strTrace) + "\n\n\n"  #see https://docs.python.org/2/library/traceback.html
        #print(Content3)
        ErrorLogFile.write(Content1)
        ErrorLogFile.write(Content2)
        ErrorLogFile.write(Content3)
        ErrorLogFile.close()
        return ErrorLog

    
              
def MergeMultipleSqliteDatabaseIntoOne(mylock,Tile, ManagementHeader,Management,Year):
    try:
        ##https://stackoverflow.com/questions/30292367/sqlite-append-two-tables-from-two-databases-that-have-the-exact-same-schema
        ##https://stackoverflow.com/questions/80801/how-can-i-merge-many-sqlite-databases
        ##https://stackoverflow.com/questions/63582979/merge-all-tables-from-sqlite3-database-into-one-single-pandas-dataframe
        ##https://www.geeksforgeeks.org/how-to-merge-multiple-csv-files-into-a-single-pandas-dataframe/
        ##https://datacarpentry.org/python-ecology-lesson/09-working-with-sql/index.html
        ##https://stackoverflow.com/questions/11653267/merge-tables-from-two-different-databases-sqlite3-python
        F3MergedDBLog = FVSprocessingpath + os.sep + "FvsDbReadyAsF3Input" + os.sep + Tile + os.sep +ManagementHeader + "_" + Management + "_"+str(Year)+"_Log.txt"
        print("We need some sentence to select the correct .DB here",file=open(F3MergedDBLog, 'a'))
        print("Step 1: from remeasured plot, choose those whose remeasuredyear is before FVS year;",file=open(F3MergedDBLog, 'a'))
        print("Step 2: if a plot has more than one remeasured, then the closest one (to FVS year) should be chosen;",file=open(F3MergedDBLog, 'a'))
        print("Step 3: from tradition or disturbed plot, remove the chosen plot",file=open(F3MergedDBLog, 'a'))
        print("Step 4: List the correct DB files for subsequent analysis by adding selected TraditionalDB, selected DisturbedFVSDB, and RemeasuredDBYearsNew_Selected_List together","\n",file=open(F3MergedDBLog, 'a'))  
        TraditionalDBfiles = glob.glob(FVSprocessingpath + os.sep + Management + os.sep + Tile+os.sep+ManagementHeader+"_"+Management+"_*_Traditional.db") #20230509 revised from AllSqliteDBfiles = glob.glob(os.getcwd()+os.sep+Management+os.sep+"*_*.db")
        TraditionalDBfiles = [k for k in TraditionalDBfiles if len(k.split(os.sep)[-1].split("_")[3]) == 20]  #revised on 20240124 from 11 to 20: assumption here is that all the FIA plot and time is 20 digits (2 unit + 4 state + 3 county + 5 plotid + 2 month + 4 year). This separate the following merged DB which has 4 digits for the Year
        TraditionalDBPlots = [k.split(os.sep)[-1].split("_")[3][0:14] for k in TraditionalDBfiles]  #revised on 20240124 from 0:5 to 0:14
        TraditionalDBYears = [k.split(os.sep)[-1].split("_")[3][16:20] for k in TraditionalDBfiles]  #revised on 20240124 from 7:11 to 16:20
        TraditionalDBPlotAndYears = [k.split(os.sep)[-1].split("_")[3] for k in TraditionalDBfiles]
        print("For FVS year=",Year," TraditionalDBPlots=",TraditionalDBPlots,file=open(F3MergedDBLog, 'a'))
        print("For FVS year=",Year," TraditionalDBYears=",TraditionalDBYears,file=open(F3MergedDBLog, 'a'))
        print("For FVS year=",Year," TraditionalDBPlotAndYears=",TraditionalDBPlotAndYears,"\n",file=open(F3MergedDBLog, 'a'))
        DisturbedFVSDBfiles = glob.glob(FVSprocessingpath + os.sep + Management + os.sep + Tile+os.sep+ManagementHeader+"_"+Management+"_*_DisturbedFVS.db") #20230509 revised from AllSqliteDBfiles = glob.glob(os.getcwd()+os.sep+Management+os.sep+"*_*.db")
        DisturbedFVSDBfiles = [k for k in DisturbedFVSDBfiles if len(k.split(os.sep)[-1].split("_")[3]) == 20]  #revised on 20240124: assumption here is that all the FIA plot and time is 20 digits (2 unit + 4 state + 3 county + 5 plotid + 2 month + 4 year). This separate the following merged DB which has 4 digits for the Year 
        DisturbedFVSDBPlots = [k.split(os.sep)[-1].split("_")[3][0:14] for k in DisturbedFVSDBfiles]
        DisturbedFVSDBYears = [k.split(os.sep)[-1].split("_")[3][16:20] for k in DisturbedFVSDBfiles]
        DisturbedFVSDBPlotAndYears = [k.split(os.sep)[-1].split("_")[3] for k in DisturbedFVSDBfiles]
        print("For FVS year=",Year," DisturbedFVSDBPlots=",DisturbedFVSDBPlots,file=open(F3MergedDBLog, 'a'))
        print("For FVS year=",Year," DisturbedFVSDBYears=",DisturbedFVSDBYears,file=open(F3MergedDBLog, 'a'))
        print("For FVS year=",Year," DisturbedFVSDBPlotAndYears=",DisturbedFVSDBPlotAndYears,"\n",file=open(F3MergedDBLog, 'a'))
        RemeasuredDBfiles = glob.glob(FVSprocessingpath + os.sep + Management + os.sep + Tile+ManagementHeader+"_"+Management+"_*_Remeasured.db") #20230509 revised from AllSqliteDBfiles = glob.glob(os.getcwd()+os.sep+Management+os.sep+"*_*.db")
        RemeasuredDBfiles = [k for k in RemeasuredDBfiles if len(k.split(os.sep)[-1].split("_")[3]) == 20]  #assumption here is that all the FIA plot ID is 5 digits. This separate the following merged DB which has 4 digits for the Year 
        RemeasuredDBPlots = [k.split(os.sep)[-1].split("_")[3][0:14] for k in RemeasuredDBfiles]
        RemeasuredDBYears = [k.split(os.sep)[-1].split("_")[3][16:20] for k in RemeasuredDBfiles]
        RemeasuredDBPlotAndYears = [k.split(os.sep)[-1].split("_")[3] for k in RemeasuredDBfiles]
        print("For FVS year=",Year," RemeasuredDBPlots=",RemeasuredDBPlots,file=open(F3MergedDBLog, 'a'))
        print("For FVS year=",Year," RemeasuredDBYears=",RemeasuredDBYears,file=open(F3MergedDBLog, 'a'))
        print("For FVS year=",Year," RemeasuredDBPlotAndYears=",RemeasuredDBPlotAndYears,"\n",file=open(F3MergedDBLog, 'a'))

        #Step 1: from remeasured plot, choose those whose remeasuredyear is before FVS year---start
        RemeasuredDBYearsNew000 = [RemeasuredDBYears[k] for k in range(0,len(RemeasuredDBYears),1) if int(RemeasuredDBYears[k]) < int(Year)]
        RemeasuredDBPlotsNew000 = [RemeasuredDBPlots[k] for k in range(0,len(RemeasuredDBPlots),1) if int(RemeasuredDBYears[k]) < int(Year)]
        RemeasuredDBPlotAndYearsNew000 = [RemeasuredDBPlotAndYears[k] for k in range(0,len(RemeasuredDBPlotAndYears),1) if int(RemeasuredDBYears[k]) < int(Year)]
        print("For FVS year=",Year," RemeasuredDBYearsNew000=",RemeasuredDBYearsNew000,file=open(F3MergedDBLog, 'a'))
        print("For FVS year=",Year," RemeasuredDBPlotsNew000=",RemeasuredDBPlotsNew000,file=open(F3MergedDBLog, 'a'))
        print("For FVS year=",Year," RemeasuredDBPlotAndYearsNew000=",RemeasuredDBPlotAndYearsNew000,file=open(F3MergedDBLog, 'a'))
        print("Added on 20230622: Remeasured Plot Must already be in tradition or disturbed, we have taken care of this during remeasured plot processing. However, the FVS input database sometimes caused problem (maybe due to missing data? For example, 59358102003 cannot be found)")
        print("Therefore, we added three lines below to solve the problem. If we do not do this, we cannot make sure that the number of plots are the identical for the same RSY")
        RemeasuredDBYearsNew = [RemeasuredDBYearsNew000[k] for k in range(0,len(RemeasuredDBYearsNew000),1) if ((RemeasuredDBPlotsNew000[k] in TraditionalDBPlots) or (RemeasuredDBPlotsNew000[k] in DisturbedFVSDBPlots))]
        RemeasuredDBPlotsNew = [RemeasuredDBPlotsNew000[k] for k in range(0,len(RemeasuredDBPlotsNew000),1) if ((RemeasuredDBPlotsNew000[k] in TraditionalDBPlots) or (RemeasuredDBPlotsNew000[k] in DisturbedFVSDBPlots))]
        RemeasuredDBPlotAndYearsNew = [RemeasuredDBPlotAndYearsNew000[k] for k in range(0,len(RemeasuredDBPlotAndYearsNew000),1) if ((RemeasuredDBPlotsNew000[k] in TraditionalDBPlots) or (RemeasuredDBPlotsNew000[k] in DisturbedFVSDBPlots))]
        print("For FVS year=",Year," RemeasuredDBYearsNew=",RemeasuredDBYearsNew,file=open(F3MergedDBLog, 'a'))
        print("For FVS year=",Year," RemeasuredDBPlotsNew=",RemeasuredDBPlotsNew,file=open(F3MergedDBLog, 'a'))
        print("For FVS year=",Year," RemeasuredDBPlotAndYearsNew=",RemeasuredDBPlotAndYearsNew,"\n",file=open(F3MergedDBLog, 'a'))
        #Step 1: from remeasured plot, choose those whose remeasuredyear is before FVS year---end

        #Step 2: if a plot has more than one remeasured, then the closest one (to FVS year) should be chosen---start
        RemeasuredDBPlotsNew_Selected_List = []
        RemeasuredDBYearsNew_Selected_List = []
        RemeasuredDBPlotAndYearsNew_Selected_List = []
        RemeasuredDBPlotsNewUnique = sorted(set(RemeasuredDBPlotsNew))  #unique and sort, see https://stackoverflow.com/questions/2931672/what-is-the-cleanest-way-to-do-a-sort-plus-uniq-on-a-python-list
        for m in RemeasuredDBPlotsNewUnique:
            m_indices  = [index for (index, item) in enumerate(RemeasuredDBPlotsNew) if item == m]  #https://www.freecodecamp.org/news/python-find-in-list-how-to-find-the-index-of-an-item-or-element-in-a-list/
            YearsOffset_m_indices = [abs(int(RemeasuredDBYearsNew[n])-int(Year)) for n in m_indices]
            YearsOffset_m_indices_minimum = min(YearsOffset_m_indices)
            RemeasuredDBPlotsNew_Selected = RemeasuredDBPlotsNew[m_indices[YearsOffset_m_indices.index(YearsOffset_m_indices_minimum)]]
            RemeasuredDBYearsNew_Selected = RemeasuredDBYearsNew[m_indices[YearsOffset_m_indices.index(YearsOffset_m_indices_minimum)]]
            RemeasuredDBPlotAndYearsNew_Selected = RemeasuredDBPlotAndYearsNew[m_indices[YearsOffset_m_indices.index(YearsOffset_m_indices_minimum)]]
            RemeasuredDBPlotsNew_Selected_List.append(RemeasuredDBPlotsNew_Selected)
            RemeasuredDBYearsNew_Selected_List.append(RemeasuredDBYearsNew_Selected)
            RemeasuredDBPlotAndYearsNew_Selected_List.append(RemeasuredDBPlotAndYearsNew_Selected)
            print("For RemeasuredDBPlotsNew=",m,"RemeasuredDBPlotsNew_Selected=",RemeasuredDBPlotsNew_Selected, " and its RemeasuredDBYearsNew_Selected=",RemeasuredDBYearsNew_Selected,file=open(F3MergedDBLog, 'a'))
        #print("RemeasuredDBPlotsNew_Selected_List=",RemeasuredDBPlotsNew_Selected_List)
        print("RemeasuredDBYearsNew_Selected_List=",RemeasuredDBYearsNew_Selected_List,"\n",file=open(F3MergedDBLog, 'a'))
        #Step 2: if a plot has more than one remeasured, then the closest one (to FVS year) should be chosen---end

        #Step 3: from tradition or disturbed plot, remove the chosen plot---start
        TraditionalDBPlots_New_Selected_List = [TraditionalDBPlots[k] for k in range(0,len(TraditionalDBPlots),1) if TraditionalDBPlots[k] not in RemeasuredDBPlotsNew_Selected_List]
        TraditionalDBYears_New_Selected_List = [TraditionalDBYears[k] for k in range(0,len(TraditionalDBPlots),1) if TraditionalDBPlots[k] not in RemeasuredDBPlotsNew_Selected_List]
        TraditionalDBPlotAndYears_New_Selected_List = [TraditionalDBPlotAndYears[k] for k in range(0,len(TraditionalDBPlots),1) if TraditionalDBPlots[k] not in RemeasuredDBPlotsNew_Selected_List]
        DisturbedFVSDBPlots_New_Selected_List = [DisturbedFVSDBPlots[k] for k in range(0,len(DisturbedFVSDBPlots),1) if DisturbedFVSDBPlots[k] not in RemeasuredDBPlotsNew_Selected_List]
        DisturbedFVSDBYears_New_Selected_List = [DisturbedFVSDBYears[k] for k in range(0,len(DisturbedFVSDBPlots),1) if DisturbedFVSDBPlots[k] not in RemeasuredDBPlotsNew_Selected_List]
        DisturbedFVSDBPlotAndYears_New_Selected_List = [DisturbedFVSDBPlotAndYears[k] for k in range(0,len(DisturbedFVSDBPlots),1) if DisturbedFVSDBPlots[k] not in RemeasuredDBPlotsNew_Selected_List]
        #Step 3: from tradition or disturbed plot, remove the chosen plot---end

        #Step 4: List the correct DB files for subsequent analysis by adding selected TraditionalDB, selected DisturbedFVSDB, and RemeasuredDBYearsNew_Selected_List together---start
        print("Here TraditionalDBPlotAndYears_New_Selected_List=",TraditionalDBPlotAndYears_New_Selected_List,file=open(F3MergedDBLog, 'a'))
        TraditionalDB = [FVSprocessingpath + os.sep + Management + os.sep + Tile+os.sep+ManagementHeader+"_"+Management+"_"+TraditionalDBPlotAndYears_New_Selected_List[k]+"_Traditional.db" for k in range(0,len(TraditionalDBPlots_New_Selected_List),1)]
        print("Here DisturbedFVSDBPlotAndYears_New_Selected_List=",DisturbedFVSDBPlotAndYears_New_Selected_List,file=open(F3MergedDBLog, 'a'))
        DisturbedFVSDB = [FVSprocessingpath + os.sep + Management + os.sep + Tile+os.sep+ManagementHeader+"_"+Management+"_"+DisturbedFVSDBPlotAndYears_New_Selected_List[k]+"_DisturbedFVS.db" for k in range(0,len(DisturbedFVSDBPlotAndYears_New_Selected_List),1)]
        print("Here RemeasuredDBYearsNew_Selected_List=",RemeasuredDBYearsNew_Selected_List,"\n",file=open(F3MergedDBLog, 'a'))
        RemeasuredDB = [FVSprocessingpath + os.sep + Management + os.sep + Tile+os.sep+ManagementHeader+"_"+Management+"_"+RemeasuredDBPlotAndYearsNew_Selected_List[k]+"_Remeasured.db" for k in range(0,len(RemeasuredDBPlotAndYearsNew_Selected_List),1)]
        AllSqliteDBfiles = TraditionalDB + DisturbedFVSDB + RemeasuredDB        
        print("AllSqliteDBfiles=",AllSqliteDBfiles,file=open(F3MergedDBLog, 'a'))
        #print("AllSqliteDBfiles=",AllSqliteDBfiles)
        #Step 4: List the correct DB files for subsequent analysis by adding selected TraditionalDB, selected DisturbedFVSDB, and RemeasuredDBYearsNew_Selected_List together---end


        MergedDB = FVSprocessingpath + os.sep + "FvsDbReadyAsF3Input" + os.sep + Tile + os.sep +ManagementHeader + "_" + Management + "_"+str(Year)+".db"
        if os.path.exists(MergedDB):
            os.remove(MergedDB) 

        #We use the first AllSqliteDBfiles to copy the structure---start
        MainConn = sqlite3.connect(MergedDB)
        MainCursor = MainConn.cursor()
        #print("AllSqliteDBfiles[0]=",AllSqliteDBfiles[0])
        conn_src = sqlite3.connect(AllSqliteDBfiles[0])  #Using the first one to copy the sqlite database structure
        conn_srcCursor = conn_src.cursor()
        for line in conn_src.iterdump():  #This is to copy the structure but do not copy the data. The code comes from ChatGPT on 20230404
            if 'INSERT INTO' not in line:  #line either "CREATE TABLE" (which is copy table and field) or "INSERT INTO" (which is record)
                MainCursor.execute(line)
        conn_src.close()
        #We use the first AllSqliteDBfiles to copy the structure---end

        #the first AllSqliteDBfile may not have al the tables, so this section search each individual SqliteDB. If a new table is found, add to the structure---start
        for IndividualDB in AllSqliteDBfiles[1:]:
            MainCursor.execute("SELECT name FROM sqlite_master WHERE type='table' ORDER BY name;")  #List all the tables in the sqlite database
            MainCursor_tables = MainCursor.fetchall()   #https://www.geeksforgeeks.org/how-to-list-tables-using-sqlite3-in-python/ and https://stackoverflow.com/questions/31986520/show-tables-in-sqlite-database-in-python
            MainCursor_tables = [k[0] for k in MainCursor_tables]  #IndividualDB_tables= [('FVS_Cases',), ('FVS_Summary2',), ('FVS_TreeList',)] so now IndividualDB_tables= ['FVS_Cases', 'FVS_Summary2', 'FVS_TreeList']

            IndividualConn = sqlite3.connect(IndividualDB)
            IndividualCursor = IndividualConn.cursor()
            IndividualCursor.execute("SELECT name FROM sqlite_master WHERE type='table' ORDER BY name;")  #List all the tables in the sqlite database
            IndividualDB_tables = IndividualCursor.fetchall()   #https://www.geeksforgeeks.org/how-to-list-tables-using-sqlite3-in-python/ and https://stackoverflow.com/questions/31986520/show-tables-in-sqlite-database-in-python
            IndividualDB_tables = [k[0] for k in IndividualDB_tables]  #IndividualDB_tables= [('FVS_Cases',), ('FVS_Summary2',), ('FVS_TreeList',)] so now IndividualDB_tables= ['FVS_Cases', 'FVS_Summary2', 'FVS_TreeList']

            MissingTablesInMainCursor = [m for m in IndividualDB_tables if m not in MainCursor_tables]
            if len(MissingTablesInMainCursor) >= 1:
                for line in IndividualConn.iterdump():  #This is to copy the structure but do not copy the data. The code comes from ChatGPT on 20230404
                    for ThisMissingTable in MissingTablesInMainCursor:
                        if ('INSERT INTO' not in line) and (ThisMissingTable in line):
                            print("New Table added with:",line)
                            print("New Table added with:",line,file=open(F3Log, 'a'))
                            MainCursor.execute(line)
            IndividualCursor.close()
       #the first AllSqliteDBfile may not have al the tables, so this section search each individual SqliteDB. If a new table is found, add to the structure---end
     
        
        
        for IndividualDB in AllSqliteDBfiles:
            IndividualConn = sqlite3.connect(IndividualDB)
            IndividualCursor = IndividualConn.cursor()
            IndividualCursor.execute("SELECT name FROM sqlite_master WHERE type='table' ORDER BY name;")  #List all the tables in the sqlite database
            IndividualDB_tables = IndividualCursor.fetchall()   #https://www.geeksforgeeks.org/how-to-list-tables-using-sqlite3-in-python/ and https://stackoverflow.com/questions/31986520/show-tables-in-sqlite-database-in-python
            IndividualDB_tables = [k[0] for k in IndividualDB_tables]  #IndividualDB_tables= [('FVS_Cases',), ('FVS_Summary2',), ('FVS_TreeList',)] so now IndividualDB_tables= ['FVS_Cases', 'FVS_Summary2', 'FVS_TreeList']

            #This part get the inventory year of the plot. Note: 1) It must match the Key part of SPLabel, 2) Only one plot in this .db, 3)FVS_TreeList table is often blank (because no tree), so we use this information---start
            if "FVS_Cases" not in IndividualDB_tables:
                print(IndividualDB," for ",ManagementHeader,Management,Year," has no FVS_Cases table, which is wrong. The reason may be the data structure. Please check",file=open(F3LogWrongFVS, 'a'))
                continue  #20230508, I found plot 99995 measured in 1994 does not have correct FVSoutputDB (I guess maybe the data structure are different).In this case, we just ignore.
            IndividualCursor.execute('SELECT * FROM FVS_Cases')   #20230508, please it is FVS_Cases rather than FVS_cases
            columns = [column[0] for column in IndividualCursor.description]
            GroupsIndex= columns.index("Groups")  #Important to make sure that "FIA_Inv_Yr=" is in the Key Template SPLabel part, which is reflected in "Groups"
            IndividualData = IndividualCursor.fetchall()
            for row in IndividualData:  #IndividualData is supposed to have only one row
                #print("row=",row)
                RowGroupContent = row[GroupsIndex]
                #print("RowGroupContent=",RowGroupContent)
                FIA_Inv_Yr = RowGroupContent.replace("FIA_Inv_Yr=","")  
            ThisPlotInventoryYear = int(FIA_Inv_Yr)
            #print("For ",IndividualDB, "ThisPlotInventoryYear=",ThisPlotInventoryYear)
            #This part get the inventory year of the plot. Note: 1) It must match the Key part of SPLabel, 2) Only one plot in this .db, 3)FVS_TreeList table is often blank (because no tree), so we use this information---end
            
            for Table in IndividualDB_tables:
                TableQuery = 'SELECT * FROM ' + Table
                #print(IndividualDB," query :", TableQuery)
                print(IndividualDB," query :", TableQuery,file=open(F3Log, 'a'))
                IndividualCursor.execute(TableQuery)   #Here we can check if a table have "Year. If yes, only select the Year; if not, then copy the whole table?
                columns = [column[0] for column in IndividualCursor.description]
                #print(IndividualDB, " table", Table," has columns of ",columns)
                print(IndividualDB, " table", Table," has columns of ",columns,file=open(F3Log, 'a'))
                if "Year" in columns:
                    YearIndex= columns.index("Year")  #also need 1) handle the empty table (e.g. no tree); 2) check the Gap year (i.e., not equal to the FVS year)
                    IndividualData = IndividualCursor.fetchall()   # Returns the results as a list.

                    #This part will select the record where row[YearIndex] = Year first. If not, then select the inventory year, which is ThisPlotInventoryYear---start
                    AllYears = []
                    for row in IndividualData:
                        AllYears.append(int(row[YearIndex]))
                    AllYears = sorted(set(AllYears))  #unique and sort, see https://stackoverflow.com/questions/2931672/what-is-the-cleanest-way-to-do-a-sort-plus-uniq-on-a-python-list
                    #print("IndividualDB,Table,AllYears=",IndividualDB,Table,AllYears)
                    if ((Year in AllYears) or (len(AllYears) == 0) or (Year >= ThisPlotInventoryYear)):  #(len(AllYears) == 0) is used here because some table is empty (e.g., FVS_TreeList may be empty due to no tree in the plot)
                        YearSelected = Year
                    else:
                        YearSelected = ThisPlotInventoryYear  #The main reason is that ThisPlotInventoryYear may be later than the FVS year (e.g., a plot measured in 2013 will not have FVS year of 2012). Note FVS always output ThisPlotInventoryYear's simulation  
                    #print("YearSelected=",YearSelected)
                    #This part will select the record where row[YearIndex] = Year first. If not, then select the inventory year, which is ThisPlotInventoryYear---end
                       
                    for row in IndividualData:
                        if str(row[YearIndex]) == str(YearSelected):
                            INSERT_INTO_Sentence1 = 'INSERT INTO '+ Table +' VALUES '
                            INSERT_INTO_Sentence2 = '(' + ",".join(['?'] * len(columns)) + ')'
                            INSERT_INTO_Sentence = INSERT_INTO_Sentence1 + INSERT_INTO_Sentence2
                            #print("INSERT_INTO_Sentence=",INSERT_INTO_Sentence)
                            #print("INSERT_INTO_Sentence=",INSERT_INTO_Sentence,file=open(F3Log, 'a'))
                            MainCursor.execute(INSERT_INTO_Sentence, row)
                if "Year" not in columns:
                    IndividualData = IndividualCursor.fetchall()   # Returns the results as a list.
                    for row in IndividualData:
                        INSERT_INTO_Sentence1 = 'INSERT INTO '+ Table +' VALUES '
                        INSERT_INTO_Sentence2 = '(' + ",".join(['?'] * len(columns)) + ')'
                        INSERT_INTO_Sentence = INSERT_INTO_Sentence1 + INSERT_INTO_Sentence2
                        #print("INSERT_INTO_Sentence=",INSERT_INTO_Sentence)
                        MainCursor.execute(INSERT_INTO_Sentence, row)
            IndividualCursor.close()
        MainConn.commit()
        MainCursor.close()


        #20230612: FVS_Cases, FVS_Summary, FVS_Compute, and mistletoe (FVS_DM_Stnd_Sum,FVS_DM_Spp_Sum) are always produced from FVS, so here I want to use FVS_Compute table to make sure the number of records are the same---start
        conn = sqlite3.connect(MergedDB)
        cursor = conn.cursor()
        cursor.execute("select * from FVS_Compute ORDER BY FIADB_PLOT ASC")   #https://www.geeksforgeeks.org/python-sqlite-order-by-clause/
        rows = cursor.fetchall()
        NumberOfRecords = len(rows)
        print("\nFor ",MergedDB," NumberOfRecords in FVS_Compute tabe =",NumberOfRecords,". If the value are different for the same RSY, there is a problem, please check!",file=open(F3MergedDBLog, 'a'))
        print("\nFor ",MergedDB," NumberOfRecords in FVS_Compute tabe =",NumberOfRecords,". If the value are different for the same RSY, there is a problem, please check!",file=open(F3Log, 'a'))
        columns = [column[0] for column in cursor.description]
        FIADB_PLOTIndex = columns.index("FIADB_PLOT")
        
        #20240213: calculate STDAGE accorind to SLTCSB and FVS STANDAGE (if SLTSCB=0, then STDAGE=0; otherise STDAGE=STANDAGE)----start
        sql_update_query0 = """Update FVS_Compute set STDAGE = 0 where SLTSCB = 0"""
        cursor.execute(sql_update_query0)
        sql_update_query1 = """Update FVS_Compute set STDAGE = STANDAGE where SLTSCB > 0"""
        cursor.execute(sql_update_query1)
        conn.commit()
        #cursor.close()  #20240318, I had an error [Sorry, there is an error :  Cannot operate on a closed cursor.], so comment it out
        #20240213: calculate STDAGE accorind to SLTCSB and FVS STANDAGE (if SLTSCB=0, then STDAGE=0; otherise STDAGE=STANDAGE)----end

        #20240229: calculate LT0AND1_0_999 accordind to SLTCSB (if SLTSCB=0, then LT0AND1_0_999=0; otherise LT0AND1_0_999=0)----start
        sql_update_query0 = """Update FVS_Compute set LT0AND1_0_999 = 0 where SLTSCB = 0"""
        cursor.execute(sql_update_query0)
        sql_update_query1 = """Update FVS_Compute set LT0AND1_0_999 = 1 where SLTSCB > 0"""
        cursor.execute(sql_update_query1)
        conn.commit()
        cursor.close()
        #20240229: calculate LT0AND1_0_999 accordind to SLTCSB (if SLTSCB=0, then LT0AND1_0_999=0; otherise LT0AND1_0_999=0)----end
        
        FIADB_PLOTList = []
        rowid = 0
        for row in rows:
            rowid = rowid + 1
            FIADB_PLOT = str(row[FIADB_PLOTIndex])
            FIADB_PLOTList.append(FIADB_PLOT)
        print(FIADB_PLOTList,file=open(F3MergedDBLog, 'a'))
        print("20240214: Please note the merge DB may have many null values. It is OK to replace these null values with zero. This has been taken care of in imputation procedure")
        #print("20240126: We need add something here or another separate function to check the NumberOfRecords for warning. It is not urgent and can be done later")
        #20230612: FVS_Cases, FVS_Summary, FVS_Compute, and mistletoe (FVS_DM_Stnd_Sum,FVS_DM_Spp_Sum) are always produced from FVS, so here I want to use FVS_Compute table to make sure the number of records are the same---end
    except Exception as e:
        print("Sorry, there is an error : ",e)
        exc_type, exc_value, exc_traceback = sys.exc_info()
        strTrace = traceback.format_exception(exc_type, exc_value, exc_traceback)
        ErrorLog = os.getcwd()+os.sep+"FvsError.txt"
        ErrorLogFile = open(ErrorLog, 'a')
        Content1 = "Dr. Shengli Huang detected an error at "+datetime.datetime.now().strftime("%Y-%m-%d %H:%M:%S") + "\n" 
        #print(Content1)
        Content2 = "The error was for MergeMultipleSqliteDatabaseIntoOne with inputs of " + repr(Tile)+","+repr(ManagementHeader)+","+repr(Management)+","+repr(Management) + " with the following details:\n"
        #print(Content2)
        Content3 = repr(strTrace) + "\n\n\n"  #see https://docs.python.org/2/library/traceback.html
        #print(Content3)
        ErrorLogFile.write(Content1)
        ErrorLogFile.write(Content2)
        ErrorLogFile.write(Content3)
        ErrorLogFile.close()
        return ErrorLog



   

if __name__ == '__main__':   
    print("\n\n\nWe cannot run in terminal; we must run under DOS command, because it is a parallel python!")
    NumberOfProcessors = int(os.cpu_count())  #os.environ["NUMBER_OF_PROCESSORS"] does not work for unix but work for windows
    print("NumberOfProcessors=",NumberOfProcessors)  #My 89 computer has 96
    UserDefinedRatio = 0.9
    #20240124: I found the error "ValueError: [Exception in thread Thread-1 (_handle_workers): need at most 63 handles, got a sequence of length 88"]
    #https://stackoverflow.com/questions/65252807/multiprocessing-pool-pool-on-windows-cpu-limit-of-63
    UserDefinedMaximumProcessors = min(60,int(NumberOfProcessors * UserDefinedRatio))  #20240124: so I use the 60 here to solve the problem above
    print("We only use ",UserDefinedMaximumProcessors," processors. The maximum value of 60 is used to solve the problem of [ValueError: [Exception in thread Thread-1 (_handle_workers): need at most 63 handles, got a sequence of length 88]")
    mymanager = multiprocessing.Manager()
    mylock = mymanager.Lock() # is "lock = multiprocessing.Lock()" OK? I guess it may not work, see https://www.thecodingforums.com/threads/multiprocessing-and-locks.679436/ and http://stackoverflow.com/questions/25557686/python-sharing-a-lock-between-processes
    myqueue = mymanager.Queue() #I do not use here, but want to keep here for future use. But rememebr this may be required for multiprocessing internally     

    for Tile in Tiles:
        GeneralKcpKeyForTile = FVSprocessingpath + os.sep + "GeneralKcpKey" + os.sep + Tile
        if not os.path.exists(GeneralKcpKeyForTile):
            os.mkdir(GeneralKcpKeyForTile)
        print(Tile, "start")
        FVS_Years = [*range(2023,2024,1)] + [*range(2028,2080,5)] #https://www.geeksforgeeks.org/range-to-a-list-in-python/
        FVS_Years = sorted(set(FVS_Years))  #unique and sort, see https://stackoverflow.com/questions/2931672/what-is-the-cleanest-way-to-do-a-sort-plus-uniq-on-a-python-list
        
        print("FVS_Years=",FVS_Years)
        FVSBatFileTemplate = FVSprocessingpath + os.sep + "GeneralKcpKey" + os.sep + "FVSBatFileTemplate.bat"
        ShengliHuangKeyFile = r'D:\CUI\subplots\L48' + os.sep + Tile + ".hsl"
        plotlist,xlist,ylist,StatesCodeInThisHSL,StatesFullNameInThisHSL,StatesShortNameInThisHSL,SpeciesTranslatorInThisHSL = PlotIDandXandY(ShengliHuangKeyFile,Factor)
        
        FVSINPUTDBs = [FVSINPUTDBpath+os.sep+"SQLite_FIADB_"+k+".db" for k in StatesShortNameInThisHSL]  #https://stackoverflow.com/questions/2860153/how-do-i-get-the-parent-directory-in-python
        print("%%%%%%%%%%%%%%%%%%%%%%%%%%%PLEASE WATCH!!!!! FVSINPUTDBs=",FVSINPUTDBs)
        SelectFIAplotMeasurementHuangList = [FVSprocessingpath +os.sep+"FiaPlotScreen"+os.sep+Tile+os.sep+"RSY"+str(k)+"_MaxGap7_Traditional.csv" for k in RemoteSensingYearOfInterest]
        print("SelectFIAplotMeasurementHuangList=",SelectFIAplotMeasurementHuangList)
        for SelectFIAplotMeasurementHuang in SelectFIAplotMeasurementHuangList:
            RSY = int(SelectFIAplotMeasurementHuang.split(os.sep)[-1].split("_")[0][3:7])
            FVS_Years = [k for k in FVS_Years if k >= RSY]
            print("FVS_Years after removing the pre-RSY =",FVS_Years)
            MaxGapYear = int(SelectFIAplotMeasurementHuang.split(os.sep)[-1].split("_")[1].replace("MaxGap",""))   
            print("MaxGapYear=",MaxGapYear)
            Management = "NoMGT" #Management = "Alt2"  #"NoMGT". Please note a Management can have multiple OutputKcpList and EventKcpList
            if "_" in Management:
                print("We do not allow _ in the Management name, because this will confuse the subsequent analysis, so we remove _ from the management name below")
                Management = Management.replace("_","")
             
            EventKcpList = ["SpeciesSizeSpecficMetrics.kcp"] #["PP_GF_100_30_template.kcp", "PP_Territory_100_24_template.kcp","SpeciesSizeSpecficMetrics.kcp"]
            SpeciesSizeSpecficMetricsKcpReturned = "SpeciesSizeSpecficMetrics is not desired"   #20240113 added: Never change this sentence
            if "SpeciesSizeSpecficMetrics.kcp" in EventKcpList:
                SpeciesGroup = ["ConOak:BO,BR,CK,NP,RO,SW,WO","Pine:JP,RN"] #The first part is species group name while the second part is species list. If you do not want to define any species group, leave this list empty. 
                SpeciesChosenMethod = "UserDefined"  #Options are "AutomaticCompleteSpecies" or "UserDefined"   #as of 20240529, the FVS output DB is empty if we use AutomaticCompleteSpecies. Maybe the number of colimns is too many? Will check later 
                if SpeciesChosenMethod == "AutomaticCompleteSpecies":  #This will use al the species in the tile
                    plotlist,FIAspeciesListInThisHSL,FVSalphaCodeListInThisHSL = PlotID_FIAspeciesCode_FVSalphaSpecies(ShengliHuangKeyFile,Factor)
                    SpeciesList = FVSalphaCodeListInThisHSL
                if SpeciesChosenMethod == "UserDefined":
                    SpeciesList = ["SM","PB","WP","BF","TA","EH","JP","RN","BO","BR","CK","NC","NP","RO","SW","WO"] #["all","PP","IC"].
                    #Marla mentioned Congress is very interested in oak species, see email Wednesday, May 1, 2024 10:21 AM and https://www.congress.gov/bill/118th-congress/house-bill/5582/text
                    #Oak species in eastern variant: WO,SW,SO,NP,SK,QI,QN,OK,TO,LK,OV,BR,DP,BJ,SN,CK,WK,NK,CB,PN,WL,CO,RO,QS,DO,PO,BO,LO
                    #Oak species in LS varaint (Michigan):BO,BR,CK,NC,NP,RO,SW,WO
                    #20240528: I thought of avoiding metrci recalculation, but the computation time of FVS is quite short compared to the whole process, so I decide not to change anything
                SizeList = ["0-999"] #["0-6","6-20","20-40","40-80","80-999"]  #Each project may have different classes, so be caustious: do NOT mess up
                if "0-999" not in SizeList:  #20240124: make sure BASA is always in the list, because BASA is used for calculating 0 absence and 1 presence
                    SizeList.append("0-999")                
                SizeList = [k.replace("-","_") for k in SizeList]  #20240125: FVS does not like - in field name, so change to _ for solving the problem
                MetricList = ["TPA","BASA","QMD","SDI1933"] #["TPA","BASA","TCFVOL","MBFVOL","QMD","AHGT","CORPC","MISTLE","MCFVOL","AVGDBHGROW","SDI1933","RELDEN1984","SDI1983","RELDEN2010"]
                if "BASA" not in MetricList:  #20240124: make sure BASA is always in the list, because BASA is used for calculating 0 absence and 1 presence
                    MetricList.append("BASA")
                SpeciesSizeSpecficMetricsKcpReturned = SpeciesSizeSpecficMetricsKcp(SpeciesList, SizeList, MetricList, SpeciesGroup)
                print("SpeciesSizeSpecficMetricsKcpReturned=",SpeciesSizeSpecficMetricsKcpReturned)
            EventKcpList = [FVSprocessingpath +os.sep+"GeneralKcpKey"+os.sep+k if k != "SpeciesSizeSpecficMetrics.kcp" else FVSprocessingpath +os.sep+"GeneralKcpKey"+os.sep+Tile+os.sep+k for k in EventKcpList]  #Revised on 20240213
            print("20230213: In the future we will move all KCP files under tile to make it more clear. It uses very little space")
            OutputKcpList = ["F3USAMetric.kcp"] #["F3USAMetric20240118.kcp","Alt2_PAC_WUI_Defense.kcp"] #["SERAL_added_WS.kcp", "CA_biomass_utilization_tool_WS.kcp", "ACCL_CWHR.kcp"]
            OutputKcpList = [FVSprocessingpath+os.sep+"GeneralKcpKey"+os.sep+k for k in OutputKcpList]
            print("IT IS SUPER IMPORTANT TO MAKE SURE THE METRICS TO BE PROCESSED ARE IN NOMGT AND MGT SCENARIOS IF YOU WANT TO COMPARE THE DIFFERENCE")
            MasterKeyFileTemplates = [FVSprocessingpath + os.sep + "GeneralKcpKey" + os.sep + "KeyTemplate20231219.key"]

            #20240529 added. This section is to automatic updating the F3MetricInformation excel file---start
            SpeciesSpecificAndSizeSpecificShortName,SpeciesSpecificAndSizeSpecificFullName = SpeciesSpecificAndSizeSpecificNamePairs(SpeciesSizeSpecficMetricsKcpReturned)
            print(SpeciesSpecificAndSizeSpecificShortName)
            print(SpeciesSpecificAndSizeSpecificFullName)
            F3MetricExcelXLSXfileMessage = AutomaticUpdateF3MetricInformationExcel(F3MetricExcelXLSXfile,SpeciesSpecificAndSizeSpecificShortName,SpeciesSpecificAndSizeSpecificFullName)  #Added on 20240513
            print(F3MetricExcelXLSXfileMessage)
            #20240529 added. This section is to automatic updating the F3MetricInformation excel file---end

            ManagementPath = FVSprocessingpath + os.sep + Management + os.sep + Tile
            if not os.path.exists(ManagementPath):
                os.makedirs(ManagementPath)
            FvsDbReadyAsF3InputPath = FVSprocessingpath + os.sep + "FvsDbReadyAsF3Input" + os.sep + Tile
            if not os.path.exists(FvsDbReadyAsF3InputPath):
                os.makedirs(FvsDbReadyAsF3InputPath)

            
            #This section creates FVS Key file Bat file, and then run FVS automatically for _Traditional.csv---start
            SelectFIAplotTraditionalFile = open(SelectFIAplotMeasurementHuang, "r")
            Lines = SelectFIAplotTraditionalFile.readlines()  #readline only read one line
            SelectFIAplotTime_Traditional = [i.replace('\n','').split(",")[0] for i in Lines]  #Most of the time, there is no comma in the line
            SelectFIAplot_Traditional = [i[0:14] for i in SelectFIAplotTime_Traditional]  #20240124 revised from: SelectFIAplot_Traditional = [i[0:5] for i in SelectFIAplotTime_Traditional]        
            print("Above are SelectFIAplotTime_Traditional and SelectFIAplot_Traditional, they are different")
            #print("SelectFIAplot_Traditional=",SelectFIAplot_Traditional)
            print("SelectFIAplot_Traditional=",SelectFIAplot_Traditional,file=open(F3Log, 'a'))
            SelectFIAplotTraditionalFile.close()
            ManagementHeader = SelectFIAplotMeasurementHuang.split(os.sep)[-1].replace("_Traditional.csv","")        
            MaximalSimutaneousTask = min(len(SelectFIAplot_Traditional)*len(MasterKeyFileTemplates), UserDefinedMaximumProcessors) 
            NumberOfSimutaneousTask = max(min(NumberOfProcessors-2, MaximalSimutaneousTask),1)  
            F3pool = multiprocessing.Pool(processes=NumberOfSimutaneousTask)
            for MasterKeyFileTemplate in MasterKeyFileTemplates:
                for SelectFIAplotMeasurement in SelectFIAplotTime_Traditional:  #Here we use SelectFIAplotTime_Traditional not SelectFIAplot_Traditional
                #for SelectFIAplotMeasurement in SelectFIAplotTime_Traditional[0:10]:   #temporal for testing. Will be removed to be retored
                    print("Traditional SelectFIAplotMeasurement=",SelectFIAplotMeasurement,file=open(F3Log, 'a'))
                    #print("We are processing Traditional SelectFIAplotMeasurement=",SelectFIAplotMeasurement)
                    Category = "Traditional"
                    F3pool.apply_async(FVSKeyBatAndRun, [mylock,Tile,Category,Management,ManagementHeader,SelectFIAplotMeasurement,MasterKeyFileTemplate,FVSBatFileTemplate,FVSINPUTDBs,FVS_Years,OutputKcpList,EventKcpList,SpeciesSizeSpecficMetricsKcpReturned,SpeciesList])   
            F3pool.close()
            F3pool.join()
            print("AAA finished at ", datetime.datetime.now().strftime("%Y-%m-%d %H:%M:%S"))
            #This section creates FVS Key file Bat file, and then run FVS automatically for _Traditional.csv---end


            #This section creates FVS Key file Bat file, and then run FVS automatically for _Remeasured.csv---start
            SelectFIAplotMeasurementHuang_Remeasured = SelectFIAplotMeasurementHuang.replace("_Traditional.csv","_Remeasured.csv")
            SelectFIAplotRemeasuredFile = open(SelectFIAplotMeasurementHuang_Remeasured, "r")
            Lines = SelectFIAplotRemeasuredFile.readlines()  #readline only read one line
            SelectFIAplotTime_Remeasured = [i.replace('\n','').split(",")[0] for i in Lines]
            SelectFIAplot_Remeasured = [i[0:14] for i in SelectFIAplotTime_Remeasured]  #20240124 change [0:5] to [0:14]
            SelectFIAplotRemeasuredFile.close()
            MaximalSimutaneousTask = min(len(SelectFIAplot_Remeasured)*len(MasterKeyFileTemplates), UserDefinedMaximumProcessors) 
            NumberOfSimutaneousTask = max(min(NumberOfProcessors-2, MaximalSimutaneousTask),1)  
            F3pool = multiprocessing.Pool(processes=NumberOfSimutaneousTask)
            for MasterKeyFileTemplate in MasterKeyFileTemplates:
                for SelectFIAplotMeasurement in SelectFIAplotTime_Remeasured: #temporal for testing. Will be removed to be retored
                    Category = "Remeasured"
                    F3pool.apply_async(FVSKeyBatAndRun, [mylock,Tile,Category,Management,ManagementHeader,SelectFIAplotMeasurement,MasterKeyFileTemplate,FVSBatFileTemplate,FVSINPUTDBs,FVS_Years,OutputKcpList,EventKcpList,SpeciesSizeSpecficMetricsKcpReturned,SpeciesList])   
            F3pool.close()
            F3pool.join()
            print("BBB finished at ", datetime.datetime.now().strftime("%Y-%m-%d %H:%M:%S"))
            #This section creates FVS Key file Bat file, and then run FVS automatically for _Remeasured.csv---end


            #When run FVS for a disturbed plot only when the plot does not exist after FVSKeyBatAndRun function (i.e., unburned is always a priortiy) ---start
            SelectFIAplotMeasurementHuang_DisturbedFVS = SelectFIAplotMeasurementHuang.replace("_Traditional.csv","_DisturbedFVS.csv")
            SelectFIAplotDisturbedFVSFile = open(SelectFIAplotMeasurementHuang_DisturbedFVS, "r")
            Lines_DisturbedFVS = SelectFIAplotDisturbedFVSFile.readlines()  #readline only read one line
            SelectFIAplotTime_DisturbedFVS = [i.replace('\n','') for i in Lines_DisturbedFVS]
            SelectFIAplot_DisturbedFVS = [i.split(",")[0] for i in SelectFIAplotTime_DisturbedFVS]
            print("SelectFIAplot_DisturbedFVS=",SelectFIAplot_DisturbedFVS,file=open(F3Log, 'a'))
            SelectFIAplotTime_DisturbedFVS_SelectedIndex = [SelectFIAplotTime_DisturbedFVS.index(m) for m in SelectFIAplotTime_DisturbedFVS if m[0:14] not in SelectFIAplot_Traditional]  #20240124 change [0:5] to [0:14]
            SelectFIAplot_DisturbedFVS_Selected = [SelectFIAplot_DisturbedFVS[k] for k in SelectFIAplotTime_DisturbedFVS_SelectedIndex]
            SelectFIAplotTime_DisturbedFVS_Selected = [SelectFIAplotTime_DisturbedFVS[k] for k in SelectFIAplotTime_DisturbedFVS_SelectedIndex]
            MaximalSimutaneousTask = min(len(SelectFIAplotTime_DisturbedFVS_Selected)*len(MasterKeyFileTemplates), UserDefinedMaximumProcessors) 
            NumberOfSimutaneousTask = max(min(NumberOfProcessors-2, MaximalSimutaneousTask),1)  
            F3pool = multiprocessing.Pool(processes=NumberOfSimutaneousTask)
            for MasterKeyFileTemplate in MasterKeyFileTemplates:
                for SelectFIAplotMeasurementDisturbedFVS in SelectFIAplotTime_DisturbedFVS_Selected: #temporal for testing. Will be removed to be retored
                    Category = "DisturbedFVS"
                    F3pool.apply_async(FVSKeyBatAndRunDisturbedFVS, [mylock,Tile,Category,Management,ManagementHeader,SelectFIAplotMeasurementDisturbedFVS,MasterKeyFileTemplate,FVSBatFileTemplate,FVSINPUTDBs,FVS_Years,OutputKcpList,EventKcpList,SpeciesSizeSpecficMetricsKcpReturned])   
            F3pool.close()
            F3pool.join()
            print("CCC finished at ", datetime.datetime.now().strftime("%Y-%m-%d %H:%M:%S"))
            #When run FVS for a disturbed plot only when the plot does not exist after FVSKeyBatAndRun function (i.e., unburned is always a priortiy) ---end


            #This section create the MGT_Year.db as F3 input---start
            MaximalSimutaneousTask = min(len(MasterKeyFileTemplates)*len(FVS_Years), UserDefinedMaximumProcessors) 
            NumberOfSimutaneousTask = max(min(NumberOfProcessors-2, MaximalSimutaneousTask),1)  
            F3pool = multiprocessing.Pool(processes=NumberOfSimutaneousTask)
            for MasterKeyFileTemplate in MasterKeyFileTemplates:
                for Year in FVS_Years: #sometimes FVS_Years[0:1] for testing, in practice, remove the index  
                    F3pool.apply_async(MergeMultipleSqliteDatabaseIntoOne, [mylock,Tile,ManagementHeader,Management,Year])   
            F3pool.close()
            F3pool.join()
            print("DDD finished at ", datetime.datetime.now().strftime("%Y-%m-%d %H:%M:%S"))
            #This section create the MGT_Year.db as F3 input---end
            print("It took \t" + str(int((time.time() - t0)/60))," minutes to run FVS and then prepare the input data for F3 run")

            
